#!/usr/bin/env python3
"""Build synthetic test workbooks for the v4 converter additions."""
import openpyxl

COMPLEMENT = str.maketrans("ACGT", "TGCA")


def revcomp(s):
    return s.translate(COMPLEMENT)[::-1]


FEATURES_HEADERS = [
    "Feature Key", "Start", "End", "Gene #", "Product", "transl_table",
    "Codon Start", "Note", "Protein ID (gb only)", "db_xref (gb only)",
    "Other Qualifiers (gb only, key=value|key=value)",
    "Partial 5' end (Y/N)", "Partial 3' end (Y/N)", "Exon Group",
    "Exception", "Transl Except", "Translation Override",
]


def add_record_info_sheet(wb, title, record_info):
    ws = wb.create_sheet(title) if title in ("dummy",) else wb.create_sheet(title)
    ws.append(["Field", "Value"])
    for k, v in record_info.items():
        ws.append([k, v])
    return ws


def add_sequence_sheet(wb, title, sequence):
    ws = wb.create_sheet(title)
    ws.append(["Sequence (one chunk per cell)"])
    for i in range(0, len(sequence), 20000):
        ws.append([sequence[i:i + 20000]])
    return ws


def add_features_sheet(wb, title, feature_rows):
    ws = wb.create_sheet(title)
    ws.append(["Features (one row per feature)"])
    ws.append(FEATURES_HEADERS)
    for row in feature_rows:
        ws.append([row.get(h, "") for h in FEATURES_HEADERS])
    return ws


def make_single_workbook(path, record_info, sequence, feature_rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_record_info_sheet(wb, "Record Info", record_info)
    add_sequence_sheet(wb, "Sequence", sequence)
    add_features_sheet(wb, "Features", feature_rows)
    wb.save(path)


def make_multi_segment_workbook(path, segments):
    """segments: list of (label, record_info, sequence, feature_rows)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for label, record_info, sequence, feature_rows in segments:
        add_record_info_sheet(wb, f"Record Info ({label})", record_info)
        add_sequence_sheet(wb, f"Sequence ({label})", sequence)
        add_features_sheet(wb, f"Features ({label})", feature_rows)
    wb.save(path)


def add(cursor, seg, segs):
    start = cursor
    cursor += len(seg)
    end = cursor - 1
    segs.append(seg)
    return cursor, start, end


# ---------------------------------------------------------------------------
# Test A: wide feature-key vocabulary + product-folding on keys that don't
# take /product per the INSDC spec.
# ---------------------------------------------------------------------------
cursor = 1
segs = []
cursor, s_utr5, e_utr5 = add(cursor, "AAAAAAAAAA", segs)             # 10 nt, 5'UTR
cursor, s_cds, e_cds = add(cursor, "ATGAAACCCGGGTTTTAG", segs)        # 18 nt CDS -> MKPGF*
cursor, s_rep, e_rep = add(cursor, "GGGGCCCCGGGGCCCC", segs)          # 16 nt repeat_region (LTR-style)
cursor, s_reg, e_reg = add(cursor, "TATAAT", segs)                    # 6 nt regulatory (promoter box)
cursor, s_utr3, e_utr3 = add(cursor, "TTTTTTTTTT", segs)              # 10 nt, 3'UTR
seqA = "".join(segs)

rowsA = [
    {"Feature Key": "5'UTR", "Start": s_utr5, "End": e_utr5},
    {"Feature Key": "CDS", "Start": s_cds, "End": e_cds, "Gene #": 1, "Product": "test protein"},
    {"Feature Key": "repeat_region", "Start": s_rep, "End": e_rep,
     "Product": "should be folded into note",  # repeat_region doesn't take /product
     "Other Qualifiers (gb only, key=value|key=value)": "rpt_type=long_terminal_repeat"},
    {"Feature Key": "regulatory", "Start": s_reg, "End": e_reg,
     "Other Qualifiers (gb only, key=value|key=value)": "regulatory_class=promoter"},
    {"Feature Key": "3'UTR", "Start": s_utr3, "End": e_utr3},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testA_wide_keys.xlsx",
    {"Locus/Sequence Name": "TestWideKeys", "Definition": "Test wide feature key vocabulary.",
     "Organism": "Test virus", "Molecule Type": "genomic DNA", "Topology": "linear", "Division": "VRL"},
    seqA, rowsA,
)
print("Test A seq length", len(seqA))

# ---------------------------------------------------------------------------
# Test B: transl_except stop-codon readthrough, plus and minus strand.
# codon1=ATG(M) codon2=AAA(K) codon3=TGA(readthrough->Trp) codon4=CCC(P)
# codon5=TAA(real stop) -> expected protein MKWP
# ---------------------------------------------------------------------------
plus_cds = "ATGAAATGACCCTAA"  # 15 nt, 5 codons
cursor = 1
segs = []
cursor, s_pad, e_pad = add(cursor, "GG", segs)
cursor, s_cdsB, e_cdsB = add(cursor, plus_cds, segs)
cursor, s_pad2, e_pad2 = add(cursor, "GG", segs)
seqB_plus = "".join(segs)

rowsB_plus = [
    {"Feature Key": "CDS", "Start": s_cdsB, "End": e_cdsB, "Gene #": 1,
     "Product": "readthrough test protein", "Transl Except": "3:Trp"},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testB_readthrough_plus.xlsx",
    {"Locus/Sequence Name": "TestReadthru", "Definition": "Test stop-codon readthrough.",
     "Organism": "Test virus", "Molecule Type": "genomic RNA", "Topology": "linear", "Division": "VRL"},
    seqB_plus, rowsB_plus,
)
print("Test B (plus) seq length", len(seqB_plus), "expected protein MKWP")

L = len(seqB_plus)
seqB_minus = revcomp(seqB_plus)


def mirror(s, e, L=L):
    return L + 1 - e, L + 1 - s


m_s, m_e = mirror(s_cdsB, e_cdsB)
rowsB_minus = [
    {"Feature Key": "CDS", "Start": m_e, "End": m_s, "Gene #": 1,
     "Product": "readthrough test protein (minus strand)", "Transl Except": "3:Trp"},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testB_readthrough_minus.xlsx",
    {"Locus/Sequence Name": "TestReadthruM", "Definition": "Test stop-codon readthrough, minus strand.",
     "Organism": "Test virus", "Molecule Type": "genomic RNA", "Topology": "linear", "Division": "VRL"},
    seqB_minus, rowsB_minus,
)
print("Test B (minus) seq length", len(seqB_minus), "CDS", m_s, m_e, "expected protein MKWP")

# ---------------------------------------------------------------------------
# Test C: Exception / Translation Override (mock RNA editing)
# ---------------------------------------------------------------------------
cursor = 1
segs = []
cursor, s_pad3, e_pad3 = add(cursor, "GG", segs)
cursor, s_cdsC1, e_cdsC1 = add(cursor, "ATGAAACCCGGGTAG", segs)   # ordinary CDS region (unused directly)
cursor, s_cdsC2, e_cdsC2 = add(cursor, "ATGGATTACAAATAG", segs)  # "edited" CDS region
seqC = "".join(segs)

rowsC = [
    # No override: should skip /translation entirely, with a warning.
    {"Feature Key": "CDS", "Start": s_cdsC1, "End": e_cdsC1, "Gene #": 1,
     "Product": "edited protein (no override given)", "Exception": "RNA editing"},
    # With override: should use the override verbatim.
    {"Feature Key": "CDS", "Start": s_cdsC2, "End": e_cdsC2, "Gene #": 2,
     "Product": "edited protein (override given)", "Exception": "RNA editing",
     "Note": "mRNA is predicted to be edited; genomic sequence alone would mistranslate this CDS",
     "Translation Override": "MDYKX"},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testC_rna_editing.xlsx",
    {"Locus/Sequence Name": "TestEdited", "Definition": "Test RNA editing exception/override.",
     "Organism": "Test paramyxovirus", "Molecule Type": "genomic RNA", "Topology": "linear",
     "Division": "VRL"},
    seqC, rowsC,
)
print("Test C seq length", len(seqC))

# ---------------------------------------------------------------------------
# Test D: native multi-segment workbook (2 segments)
# ---------------------------------------------------------------------------
seg1_seq = "GG" + "ATGAAACCCGGGTTTTAG" + "GG"   # CDS -> MKPGF
seg2_seq = "CC" + "ATGGATTACAAATAG" + "CC"       # CDS -> MDYK
seg1_rows = [{"Feature Key": "CDS", "Start": 3, "End": 20, "Gene #": 1, "Product": "segment 1 protein"}]
seg2_rows = [{"Feature Key": "CDS", "Start": 3, "End": 17, "Gene #": 1, "Product": "segment 2 protein"}]

make_multi_segment_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testD_multisegment.xlsx",
    [
        ("Seg1", {"Locus/Sequence Name": "TestSeg1", "Definition": "Test segment 1.",
                   "Organism": "Test segmented virus", "Molecule Type": "genomic RNA",
                   "Topology": "linear", "Division": "VRL"}, seg1_seq, seg1_rows),
        ("Seg2", {"Locus/Sequence Name": "TestSeg2", "Definition": "Test segment 2.",
                   "Organism": "Test segmented virus", "Molecule Type": "genomic RNA",
                   "Topology": "linear", "Division": "VRL"}, seg2_seq, seg2_rows),
    ],
)
print("Test D segments: seg1 len", len(seg1_seq), "seg2 len", len(seg2_seq))
print("expected proteins: MKPGF and MDYK")

# ---------------------------------------------------------------------------
# Test E: plus-strand circular-origin-spanning CDS (e.g. geminivirus Rep-
# style wrap). Genome length 30; gene wraps from 25-30 to 1-9.
# exon1 (tail, 25-30) = ATGAAA (M K); exon2 (head, 1-9) = CCCGATTAG (P D *)
# expected protein: MKPD
# ---------------------------------------------------------------------------
exon2_wrap = "CCCGATTAG"   # positions 1-9
filler = "N" * 15          # positions 10-24
exon1_wrap = "ATGAAA"      # positions 25-30
seqE = exon2_wrap + filler + exon1_wrap
assert len(seqE) == 30, len(seqE)

rowsE = [
    {"Feature Key": "CDS", "Start": 25, "End": 30, "Gene #": 1,
     "Product": "origin-wrapping test protein", "Exon Group": "wrap1"},
    {"Feature Key": "CDS", "Start": 1, "End": 9, "Exon Group": "wrap1"},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testE_origin_wrap.xlsx",
    {"Locus/Sequence Name": "TestWrap", "Definition": "Test circular-origin-spanning CDS.",
     "Organism": "Test geminivirus", "Molecule Type": "genomic DNA", "Topology": "circular",
     "Division": "VRL"},
    seqE, rowsE,
)
print("Test E seq length", len(seqE), "expected protein MKPD")

# ---------------------------------------------------------------------------
# Test F: HBV-style overlapping reading frames (two CDS rows overlapping in
# different frames, same strand, no Exon Group involved).
# ---------------------------------------------------------------------------
seqF = "GGATGAAACCCGGGTTTAAAGGGCATTAGCCC"  # 32 nt, arbitrary
rowsF = [
    {"Feature Key": "CDS", "Start": 3, "End": 17, "Gene #": 1, "Product": "overlap frame 1"},
    {"Feature Key": "CDS", "Start": 7, "End": 24, "Gene #": 2, "Product": "overlap frame 2"},
]
make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testF_overlap.xlsx",
    {"Locus/Sequence Name": "TestOverlap", "Definition": "Test overlapping reading frames (HBV-style).",
     "Organism": "Test hepadnavirus", "Molecule Type": "genomic DNA", "Topology": "circular",
     "Division": "VRL"},
    seqF, rowsF,
)
print("Test F seq length", len(seqF))

print("\nAll v4 test workbooks written.")
