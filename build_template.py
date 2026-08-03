"""
Builds GenBank_Annotation_Template.xlsx — a non-expert-friendly Excel workbook
for biologists to enter sequence + feature annotation data.

v2: Features sheet redesigned to match the lab's real NCBI 5-column feature
table format (see Feature Moonfish_annotation_KP.txt) — the file that gets
uploaded directly to NCBI BankIt/WebSub. Two converters read this same
workbook:
  - xlsx_to_feature_table.py  -> the 5-column tab-delimited feature table
    (BankIt upload format, matches the Moonfish file's structure exactly)
  - xlsx_to_genbank.py        -> a full GenBank flat file (.gb)

Sheets:
  1. Instructions   - how to use the template, column-by-column
  2. Record Info     - key/value sheet for record-level metadata + sequence
  3. Features         - one row per feature (gene#, key, start, end, product...)
  4. Lookup Lists       (hidden) - backing lists for dropdown validation
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
OPT_HEADER_FILL = PatternFill("solid", fgColor="808080")
OPT_HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEADER_FONT = Font(name=FONT_NAME, bold=True, size=10)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # yellow-ish = fill this in
LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10)
BODY_FONT = Font(name=FONT_NAME, size=10)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ---------------------------------------------------------------------------
# Sheet: Instructions
# ---------------------------------------------------------------------------
ws = wb.active
ws.title = "Instructions"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 90

ws["B2"] = "GenBank / BankIt Annotation Template"
ws["B2"].font = TITLE_FONT
ws["B3"] = ("Fill in your sequence + feature data here. xlsx_to_feature_table.py produces the 5-column "
            "feature table for NCBI BankIt upload; xlsx_to_genbank.py produces a full GenBank flat file (.gb).")
ws["B3"].font = Font(name=FONT_NAME, size=11, italic=True)
ws["B3"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B3:F3")
ws.row_dimensions[3].height = 28

row = 5
def section(title):
    global row
    c = ws.cell(row=row, column=2, value=title)
    c.font = Font(name=FONT_NAME, bold=True, size=12, color="1F4E78")
    row += 1

def bullet(text):
    global row
    c = ws.cell(row=row, column=2, value="•")
    c.font = BODY_FONT
    c2 = ws.cell(row=row, column=3, value=text)
    c2.font = BODY_FONT
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 28 if len(text) > 90 else 15
    row += 1

section("How this workbook is organized")
bullet("\"Record Info\" sheet: record-level metadata (organism, molecule type, topology, source details, and the "
       "nucleotide sequence itself). Locus/Sequence Name doubles as the feature table's sequence ID and the prefix "
       "used for locus_tag values (e.g. \"Moonfish_gp1\").")
bullet("\"Features\" sheet: one row per gene, in genome order (top to bottom = start to end of the sequence). "
       "Blue-headed columns are the ones that go into the feature table uploaded to BankIt. Gray-headed columns "
       "are optional extras only used when generating a full GenBank flat file (.gb).")
bullet("Cells shaded pale yellow are the ones you should fill in. A few example rows are included, styled in green "
       "italics so they're not mistaken for your own data — delete or overwrite them.")

row += 1
section("Record Info sheet — field notes")
bullet("Locus/Sequence Name: a short identifier, e.g. \"Moonfish\" — used as the feature table sequence ID and the "
       "locus_tag prefix. Required.")
bullet("Definition, Organism, Molecule Type, Topology, Division: used for the full .gb flat file and as a reference "
       "when filling out the BankIt web form by hand. Required for the .gb output.")
bullet("Strain / Isolate / Collection Date / Country: source modifiers — BankIt asks for these directly in its own "
       "form (step \"Source Modifiers\"), so treat these as a reference copy, not something the scripts upload for you.")
bullet("The nucleotide sequence itself goes on the separate \"Sequence\" sheet, not here — Excel caps a single "
       "cell at 32,767 characters (the actual Excel limit), which most real genomes exceed, so that sheet lets "
       "you split the sequence across multiple cells. Required — used to sanity-check feature coordinates and "
       "(for the .gb output) to compute CDS translations.")

row += 1
section("Features sheet — column notes (matches the lab's feature-table format)")
bullet("Feature Key: CDS, tRNA, rRNA, misc_feature, or gene. For CDS rows you do NOT need a separate gene row — "
       "the script automatically writes the paired gene feature (with locus_tag) above each CDS, exactly like the "
       "real feature table does.")
bullet("Start / End: nucleotide positions (1-based). Strand is implied by the order, per NCBI convention: if Start "
       "< End the feature is on the + strand; if Start > End (a higher number first) it's on the - strand. Don't "
       "add a separate strand column — just flip the two numbers.")
bullet("Gene #: the gene number for CDS features (e.g. 1, 2, 3...) — becomes locus_tag \"{Name}_gp{N}\" and the "
       "short product label \"gp{N}\". Leave blank for tRNA/rRNA/misc_feature rows.")
bullet("Product: the descriptive protein/RNA name (e.g. \"tail spike protein\", \"tRNA-Ile\"). Required for CDS/tRNA/rRNA.")
bullet("transl_table: the NCBI translation table number. Leave blank to default to 11 (bacteria/archaea/phage — "
       "correct for essentially all phage annotation). Only used for CDS.")
bullet("Codon Start: 1, 2, or 3 — which nucleotide begins the first complete codon. Leave blank to default to 1 "
       "(standard for a complete CDS). Only used for CDS.")
bullet("Note: free-text note (e.g. \"similar to Escherichia phage phi G17\", BLAST/InterPro hits). Optional.")
bullet("Protein ID / db_xref / Other Qualifiers (gray headers): optional, only read by xlsx_to_genbank.py when "
       "building a full flat file. Other Qualifiers uses key=value pairs separated by |, e.g. "
       "function=DNA repair|EC_number=3.1.11.- — avoid using a | character within the value itself, since "
       "that's the separator between qualifiers.")
bullet("Partial 5' end / Partial 3' end (Y/N): optional, used by BOTH scripts. Set to Y if the feature is cut off "
       "(incomplete) at that end of your sequenced region — this adds the standard < or > partial marker to the "
       "feature's start/end position, regardless of strand.")

row += 1
section("Two ways to convert this workbook")
bullet("python3 xlsx_to_feature_table.py filled.xlsx moonfish.tbl.txt  → the 5-column tab-delimited feature table "
       "you upload directly in the BankIt \"Feature\" step.")
bullet("python3 xlsx_to_genbank.py filled.xlsx output.gb  → a complete GenBank flat file (LOCUS/DEFINITION/.../"
       "FEATURES/ORIGIN) for record-keeping, sharing, or non-BankIt submission routes.")

row += 1
section("Things this template does NOT support (v1)")
bullet("Multi-exon / spliced features (join(...) locations) — each feature is a single contiguous span.")
bullet("Multiple sequence records in one workbook — use one filled copy of this template per record/genome.")
bullet("A 'source' line in the feature table — BankIt collects organism/strain/collection date/etc. through its "
       "own web form, so the feature-table output intentionally has no source feature (matching real BankIt "
       "submissions). Record Info still captures these fields as your reference copy.")

row += 1
section("Reference")
bullet("NCBI Feature Table specification: https://www.ncbi.nlm.nih.gov/genbank/feature_table/")
bullet("NCBI sample GenBank record with field-by-field notes: https://www.ncbi.nlm.nih.gov/genbank/samplerecord/")
bullet("NCBI BankIt / WebSub: https://www.ncbi.nlm.nih.gov/WebSub/")

# ---------------------------------------------------------------------------
# Hidden Lookup Lists sheet (backing data for dropdowns)
# ---------------------------------------------------------------------------
lk = wb.create_sheet("Lookup Lists")
lk.sheet_state = "hidden"

feature_keys = ["CDS", "tRNA", "rRNA", "misc_feature", "gene"]
yn = ["Y", "N"]
mol_types = ["genomic DNA", "mRNA", "rRNA", "tRNA", "transcribed RNA", "other DNA", "other RNA", "viral cRNA"]
topologies = ["linear", "circular"]
divisions = [
    "PRI - primate", "ROD - rodent", "MAM - other mammalian", "VRT - other vertebrate",
    "INV - invertebrate", "PLN - plant/fungal/algal", "BCT - bacterial", "VRL - viral",
    "PHG - bacteriophage", "SYN - synthetic", "UNA - unannotated", "EST - EST",
    "PAT - patent", "STS - STS", "GSS - GSS", "HTG - high-throughput genomic",
    "HTC - high-throughput cDNA", "ENV - environmental sampling",
]
codon_starts = ["1", "2", "3"]
transl_tables = ["11", "1"]

def write_list(col_letter, name, values):
    lk[f"{col_letter}1"] = name
    for i, v in enumerate(values, start=2):
        lk[f"{col_letter}{i}"] = v
    lk.column_dimensions[col_letter].width = 28

write_list("A", "FeatureKeys", feature_keys)
write_list("B", "YN", yn)
write_list("C", "MolTypes", mol_types)
write_list("D", "Topologies", topologies)
write_list("E", "Divisions", divisions)
write_list("F", "CodonStarts", codon_starts)
write_list("G", "TranslTables", transl_tables)

def defname(name, col, n):
    wb.defined_names[name] = openpyxl.workbook.defined_name.DefinedName(
        name, attr_text=f"'Lookup Lists'!${col}$2:${col}${1+n}")

defname("FeatureKeyList", "A", len(feature_keys))
defname("YNList", "B", len(yn))
defname("MolTypeList", "C", len(mol_types))
defname("TopologyList", "D", len(topologies))
defname("DivisionList", "E", len(divisions))
defname("CodonStartList", "F", len(codon_starts))
defname("TranslTableList", "G", len(transl_tables))

# ---------------------------------------------------------------------------
# Sheet: Record Info  (key/value layout)
# ---------------------------------------------------------------------------
ri = wb.create_sheet("Record Info")
ri.sheet_view.showGridLines = False
ri.column_dimensions["A"].width = 22
ri.column_dimensions["B"].width = 55
ri.column_dimensions["C"].width = 55

ri["A1"] = "Field"
ri["B1"] = "Value  (fill in the yellow cells)"
ri["C1"] = "Example"
for col in ("A1", "B1", "C1"):
    ri[col].font = HEADER_FONT
    ri[col].fill = HEADER_FILL
    ri[col].alignment = Alignment(vertical="center")
ri.row_dimensions[1].height = 20

fields = [
    ("Locus/Sequence Name", "", "Moonfish", True),
    ("Definition", "", "Escherichia phage Moonfish, complete genome", True),
    ("Organism", "", "Escherichia virus Moonfish", True),
    ("Molecule Type", "", "genomic DNA", True),
    ("Topology", "", "linear", True),
    ("Division", "", "PHG - bacteriophage", True),
    ("Accession", "", "(leave blank if none assigned)", False),
    ("Strain", "", "", False),
    ("Isolate", "", "", False),
    ("Collection Date", "", "15-Jun-2025", False),
    ("Country", "", "USA: Massachusetts", False),
    ("Comment", "", "Annotated per dsDNA Phage Genome Sequencing and Annotation Protocol.", False),
]
field_rows = {}  # label -> row number, so dropdown wiring below never has to hardcode row numbers
r = 2
for label, val, example, required in fields:
    field_rows[label] = r
    ri.cell(row=r, column=1, value=label + (" *" if required else "")).font = LABEL_FONT
    vcell = ri.cell(row=r, column=2, value=val)
    vcell.fill = INPUT_FILL
    vcell.font = BODY_FONT
    vcell.border = BORDER
    vcell.alignment = Alignment(wrap_text=True, vertical="top")
    ecell = ri.cell(row=r, column=3, value=example)
    ecell.font = NOTE_FONT
    ecell.alignment = Alignment(wrap_text=True, vertical="top")
    ri.row_dimensions[r].height = 16
    r += 1

ri.cell(row=r + 1, column=1, value="* required field").font = NOTE_FONT
ri.cell(row=r + 2, column=1,
         value="Nucleotide sequence goes on the separate \"Sequence\" sheet, not here (see that sheet's instructions).").font = NOTE_FONT

dv_mol = DataValidation(type="list", formula1="=MolTypeList", allow_blank=True, showDropDown=False)
dv_top = DataValidation(type="list", formula1="=TopologyList", allow_blank=True, showDropDown=False)
dv_div = DataValidation(type="list", formula1="=DivisionList", allow_blank=True, showDropDown=False)
ri.add_data_validation(dv_mol)
ri.add_data_validation(dv_top)
ri.add_data_validation(dv_div)
dv_mol.add(ri[f"B{field_rows['Molecule Type']}"])
dv_top.add(ri[f"B{field_rows['Topology']}"])
dv_div.add(ri[f"B{field_rows['Division']}"])

# ---------------------------------------------------------------------------
# Sheet: Sequence
# ---------------------------------------------------------------------------
sq = wb.create_sheet("Sequence")
sq.sheet_view.showGridLines = False
sq.column_dimensions["A"].width = 100

sq["A1"] = ("Paste your nucleotide sequence below, starting in cell A2 — one continuous string per cell, "
            "no spaces/numbers/line breaks (A/C/G/T/U and standard IUPAC ambiguity codes only). Excel limits a "
            "single cell to 32,767 characters (the actual Excel limit), so for anything longer than that (most "
            "real genomes), split the sequence across multiple cells in this column (A2, A3, A4, ...) IN ORDER — "
            "the scripts read down the column and join the cells together, so where you break the sequence "
            "between cells does not matter as long as the order is right. Don't put anything else on this sheet.")
sq["A1"].font = NOTE_FONT
sq["A1"].alignment = Alignment(wrap_text=True, vertical="top")
sq.row_dimensions[1].height = 60
sq_example = PatternFill("solid", fgColor="FFF2CC")
sq["A2"].fill = sq_example
sq["A2"].font = BODY_FONT
sq.row_dimensions[2].height = 26

# ---------------------------------------------------------------------------
# Sheet: Features
# ---------------------------------------------------------------------------
fs = wb.create_sheet("Features")
fs.sheet_view.showGridLines = False
fs.freeze_panes = "A3"

core_headers = [
    "Feature Key", "Start", "End", "Gene #", "Product", "transl_table",
    "Codon Start", "Note",
]
optional_headers = [
    "Protein ID (gb only)", "db_xref (gb only)", "Other Qualifiers (gb only, key=value|key=value)",
    "Partial 5' end (Y/N)", "Partial 3' end (Y/N)",
]
headers = core_headers + optional_headers
widths = [13, 8, 8, 9, 26, 12, 12, 30, 16, 14, 34, 16, 16]

fs["A1"] = ("Fill in one row per feature, in genome order. Blue headers = used in the BankIt feature table. "
            "Gray headers = optional, only used for the full GenBank flat file. See Instructions sheet for details.")
fs["A1"].font = NOTE_FONT
fs["A1"].alignment = Alignment(wrap_text=True)
fs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
fs.row_dimensions[1].height = 28

for i, h in enumerate(headers, start=1):
    c = fs.cell(row=2, column=i, value=h)
    if i <= len(core_headers):
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    else:
        c.font = OPT_HEADER_FONT
        c.fill = OPT_HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    fs.column_dimensions[get_column_letter(i)].width = widths[i - 1]
fs.row_dimensions[2].height = 32

# Example rows, modeled on the real Moonfish feature table (genes 1-3 + the tRNA-Ile entry)
example_rows = [
    ["CDS", 39, 974, 1, "tailspike protein", "", "", "similar to gp52 in HRP29", "", "", "", "N", "N"],
    ["CDS", 1175, 1408, 2, "tail length tape-measure protein", "", "", "similar to Escherichia phage vB_Eco_F22", "", "", "", "N", "N"],
    ["CDS", 1393, 1653, 3, "putative holin", "", "", "similar to Escherichia phage IME11", "", "", "", "N", "N"],
    ["tRNA", 26894, 26821, "", "tRNA-Ile", "", "", "predicted by tRNAscan-SE", "", "", "", "N", "N"],
]

r = 3
for row_vals in example_rows:
    for i, v in enumerate(row_vals, start=1):
        c = fs.cell(row=r, column=i, value=v)
        c.font = Font(name=FONT_NAME, size=10, italic=True, color="375623")
        c.fill = PatternFill("solid", fgColor="E2EFDA")
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
    r += 1

for _ in range(40):
    for i in range(1, len(headers) + 1):
        c = fs.cell(row=r, column=i)
        c.fill = INPUT_FILL
        c.font = BODY_FONT
        c.border = BORDER
    r += 1

last_row = r - 1

dv_feat = DataValidation(type="list", formula1="=FeatureKeyList", allow_blank=True, showDropDown=False)
dv_transl = DataValidation(type="list", formula1="=TranslTableList", allow_blank=True, showDropDown=False)
dv_codon = DataValidation(type="list", formula1="=CodonStartList", allow_blank=True, showDropDown=False)
dv_yn5 = DataValidation(type="list", formula1="=YNList", allow_blank=True, showDropDown=False)
dv_yn3 = DataValidation(type="list", formula1="=YNList", allow_blank=True, showDropDown=False)
for dv in (dv_feat, dv_transl, dv_codon, dv_yn5, dv_yn3):
    fs.add_data_validation(dv)

dv_feat.add(f"A3:A{last_row}")
dv_transl.add(f"F3:F{last_row}")
dv_codon.add(f"G3:G{last_row}")
dv_yn5.add(f"L3:L{last_row}")
dv_yn3.add(f"M3:M{last_row}")

wb.save("GenBank_Annotation_Template.xlsx")
print("Workbook built: GenBank_Annotation_Template.xlsx")
