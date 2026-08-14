#!/usr/bin/env python3
"""Build a worked-example test workbook for Karen Maxwell-style phage
anti-CRISPR (acr) / anti-CRISPR-associated (aca) operon annotation.

Real acr genes are famously hard to identify from sequence alone (they're
small, fast-evolving, and share no common motif). The standard workaround
used in the field (Pawluk et al.) is "guilt by association": aca genes
(helix-turn-helix transcriptional regulators that autoregulate the acr-aca
operon) ARE recognizable by their HTH domain, and acr genes are typically
found immediately upstream of an aca gene in the same operon. Annotating
that operon relationship explicitly (feature key `operon`, `/note` guilt-by-
association rationale) is exactly the kind of structured annotation this
converter's Other Qualifiers + operon feature key support was built for.

This builds a small synthetic 2-gene operon: an acr-like CDS immediately
upstream of an aca-like CDS, wrapped in a single `operon` feature.
"""
import openpyxl

FEATURES_HEADERS = [
    "Feature Key", "Start", "End", "Gene #", "Product", "transl_table",
    "Codon Start", "Note", "Protein ID (gb only)", "db_xref (gb only)",
    "Other Qualifiers (gb only, key=value|key=value)",
    "Partial 5' end (Y/N)", "Partial 3' end (Y/N)", "Exon Group",
    "Exception", "Transl Except", "Translation Override",
]


def add_record_info_sheet(wb, title, record_info):
    ws = wb.create_sheet(title)
    ws.append(["Field", "Value"])
    for k, v in record_info.items():
        ws.append([k, v])
    return ws


def add_sequence_sheet(wb, title, sequence):
    ws = wb.create_sheet(title)
    ws.append(["Sequence (one chunk per cell)"])
    for i in range(0, len(sequence), 20000):
        ws.append([sequence[i:i + 20000]])
    return ws


def add_features_sheet(wb, title, feature_rows):
    ws = wb.create_sheet(title)
    ws.append(["Features (one row per feature)"])
    ws.append(FEATURES_HEADERS)
    for row in feature_rows:
        ws.append([row.get(h, "") for h in FEATURES_HEADERS])
    return ws


def make_single_workbook(path, record_info, sequence, feature_rows):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    add_record_info_sheet(wb, "Record Info", record_info)
    add_sequence_sheet(wb, "Sequence", sequence)
    add_features_sheet(wb, "Features", feature_rows)
    wb.save(path)


acr_cds = "ATGAAACCCTAG"      # 12 nt -> M K P (stop) -> protein "MKP"
aca_cds = "ATGGATTACAAATAG"   # 15 nt -> M D Y K (stop) -> protein "MDYK"

seq = "GG" + acr_cds + "GG" + aca_cds + "GG"
s_acr, e_acr = 3, 2 + len(acr_cds)              # 3..14
s_aca, e_aca = e_acr + 3, e_acr + 2 + len(aca_cds)  # 17..31
s_operon, e_operon = s_acr, e_aca               # 3..31

assert seq[s_acr - 1:e_acr] == acr_cds
assert seq[s_aca - 1:e_aca] == aca_cds

rows = [
    {"Feature Key": "operon", "Start": s_operon, "End": e_operon,
     "Note": "Predicted acr-aca operon, identified by guilt-by-association: "
             "acrX1 lacks a recognizable domain but sits immediately upstream "
             "of aca1, an HTH-domain autoregulator typical of anti-CRISPR "
             "operons (Pawluk et al. 2016 guilt-by-association strategy).",
     "Other Qualifiers (gb only, key=value|key=value)": "operon=acrX1-aca1 operon"},
    {"Feature Key": "CDS", "Start": s_acr, "End": e_acr, "Gene #": 1,
     "Product": "putative anti-CRISPR protein AcrX1",
     "Note": "Small ORF with no recognizable conserved domain; candidate acr "
             "gene by position immediately upstream of aca1."},
    {"Feature Key": "CDS", "Start": s_aca, "End": e_aca, "Gene #": 2,
     "Product": "anti-CRISPR-associated protein Aca1",
     "Note": "Helix-turn-helix transcriptional autoregulator; hallmark of "
             "acr-aca operons, used to flag adjacent genes as acr candidates."},
]

make_single_workbook(
    "/sessions/bold-exciting-brown/mnt/outputs/testG_acr_aca_operon.xlsx",
    {"Locus/Sequence Name": "TestAcrAca", "Definition": "Test acr-aca operon annotation.",
     "Organism": "Test phage", "Molecule Type": "genomic DNA", "Topology": "linear",
     "Division": "VRL"},
    seq, rows,
)
print("Sequence length:", len(seq))
print("operon:", s_operon, e_operon)
print("acr CDS:", s_acr, e_acr, "expected protein MKP")
print("aca CDS:", s_aca, e_aca, "expected protein MDYK")
