#!/usr/bin/env python3
"""
xlsx_to_feature_table.py — convert a filled copy of GenBank_Annotation_Template.xlsx
into NCBI's 5-column tab-delimited feature table: the exact file format uploaded
in the "Feature" step of NCBI BankIt/WebSub submissions.

Usage:
    python3 xlsx_to_feature_table.py input.xlsx [output.tbl.txt]

Format (per https://www.ncbi.nlm.nih.gov/genbank/feature_table/ and matching
the structure of a real lab feature table, Feature Moonfish_annotation_KP.txt):

    >Feature	<Locus/Sequence Name>
    <start>	<end>	gene
    			locus_tag	<Name>_gp<N>
    <start>	<end>	CDS
    			product	<descriptive product>
    			product	gp<N>
    			transl_table	11
    			codon_start	1
    			note	<note>            (only if a note was given)

Notes on conventions implemented, matching this lab's protocol:
  - Strand is NOT a separate column: if Start < End the feature is '+', if
    Start > End it's '-' — you enter the numbers in that order directly, and
    they're passed straight through to the two location columns.
  - Every CDS row automatically gets a paired 'gene' feature written above it
    (with a locus_tag, if a Gene # was given) — you do not enter a separate
    gene row for each CDS.
  - CDS features always get transl_table (default 11) and codon_start
    (default 1) qualifiers, matching standard phage annotation practice.
  - Partial features get the standard '<'/'>' markers on the start/end
    columns (column 1 gets '<', column 2 gets '>', regardless of strand).
  - There is no 'source' feature in this file — BankIt collects organism/
    strain/collection date/etc. through its own web form, not the feature
    table upload.
  - Qualifier lines always use real tabs (three leading tabs, then key, tab,
    value) and blank/empty qualifiers are simply omitted, so the output
    doesn't carry the stray blank-line and missing-tab artifacts sometimes
    seen in hand-edited feature tables. Embedded tabs/newlines inside a
    value are sanitized (collapsed to a space, with a warning) since this
    format is itself tab/newline-delimited.

v3 additions: mat_peptide feature key (polyprotein viruses) and an Exon
Group column for multi-interval (spliced) features — see the v3 changelog
retained in xlsx_to_genbank.py for the full rationale, which applies to
both scripts since they share a workbook and the same entry-order
convention.

v4 additions (broader virus/genome coverage):
  - VALID_FEATURE_KEYS now covers the full INSDC feature key vocabulary
    (confirmed against the official spec at
    https://www.insdc.org/submitting-standards/feature-table/), not just
    the handful this tool originally special-cased. New keys use a generic
    handler (Product if given, Note if given, Other Qualifiers) — the same
    treatment misc_feature always got. Two controlled-vocabulary qualifiers
    that come up often with these keys are reachable through the existing
    Other Qualifiers column rather than getting bespoke columns of their
    own: put `regulatory_class=promoter` (or terminator, ribosome_binding_
    site, etc.) in Other Qualifiers for a `regulatory` feature, and
    `rpt_type=long_terminal_repeat` (or inverted, direct, dispersed, etc.)
    for a `repeat_region` feature (this is also the modern, correct way to
    annotate retroviral LTRs and viral inverted terminal repeats — INSDC
    dropped the old standalone "LTR"/"inverted_repeat" keys in favor of
    repeat_region + rpt_type).
  - Exception and Transl Except columns, for CDS features only: Exception
    writes /exception="..." (e.g. "RNA editing" for paramyxovirus P/V/W-
    style edited transcripts, or "trans-splicing"). Transl Except writes
    one or more /transl_except=(pos:<location>,aa:<amino_acid>) qualifiers
    for single-codon deviations from the standard translation — most
    commonly a stop-codon readthrough (alphaviruses, some plant viruses)
    or selenocysteine incorporation. You give the codon NUMBER within the
    CDS (e.g. "142:Trp" for the 142nd codon read through as tryptophan);
    this script computes the correct genomic location for you. Only
    supported for non-spliced (single-interval) CDS rows — computing which
    genomic interval a codon number falls in across an intron boundary is
    not implemented, so a spliced CDS with Transl Except gets a warning and
    the qualifier is skipped rather than risking a wrong location.
  - Native multi-segment workbooks: a workbook may contain repeated sheet
    triplets named "Record Info (X)"/"Sequence (X)"/"Features (X)" (X = any
    segment label, e.g. "PB2", "Seg1") for genomes like influenza or
    bunyaviruses where each segment is its own GenBank record. Detected
    automatically; a plain "Record Info"/"Sequence"/"Features" workbook
    (no suffix) is treated as a single segment, unchanged from v3. Output
    is one combined .tbl file (repeated >Feature blocks, valid per the
    INSDC spec) plus a companion multi-FASTA (.fsa) of all segments'
    sequences, matching how BankIt batch submissions actually pair a
    multi-FASTA with a multi-record feature table.
  - Plus-strand circular-origin-spanning features: if an Exon Group's
    intervals are not in strictly ascending order for a plus-strand
    feature (e.g. a geminivirus Rep gene whose join() wraps from near the
    end of a circular sequence back to position 1 — join(2004..2195,3..20)
    is the INSDC spec's own worked example of this), the feature is
    detected as origin-spanning and its auto-generated paired gene feature
    is given the same multi-interval join() as the CDS instead of an
    incorrect single min-to-max span (which would wrongly claim the gene
    covers the entire genome in between). Minus-strand origin-spanning
    features are not supported — flagged with a warning instead of guessed
    at, since the correct INSDC representation there needs the alternate
    join(complement(...),complement(...)) syntax form, which was not
    implemented or tested here.
"""
import sys
import re
from openpyxl import load_workbook

# Full INSDC feature key vocabulary (source: official DDBJ/ENA/GenBank
# Feature Table Definition, https://www.insdc.org/submitting-standards/feature-table/).
VALID_FEATURE_KEYS = {
    "assembly_gap", "C_region", "CDS", "centromere", "D-loop", "D_segment",
    "exon", "gap", "gene", "iDNA", "intron", "J_segment", "mat_peptide",
    "misc_binding", "misc_difference", "misc_feature", "misc_recomb",
    "misc_RNA", "misc_structure", "mobile_element", "modified_base", "mRNA",
    "ncRNA", "N_region", "old_sequence", "operon", "oriT", "polyA_site",
    "precursor_RNA", "prim_transcript", "primer_bind", "propeptide",
    "protein_bind", "regulatory", "repeat_region", "rep_origin", "rRNA",
    "S_region", "sig_peptide", "stem_loop", "STS", "telomere", "tmRNA",
    "transit_peptide", "tRNA", "unsure", "V_region", "V_segment",
    "variation", "3'UTR", "5'UTR",
    # source is handled as an internal-only feature in xlsx_to_genbank.py;
    # this script never has source rows (BankIt collects that separately),
    # but accept it here too rather than reject it outright if present.
    "source",
}

# Feature keys where a /product qualifier is expected/recommended, beyond
# CDS and mat_peptide (which have their own dedicated handling below).
RNA_PRODUCT_KEYS = {"tRNA", "rRNA", "tmRNA", "ncRNA", "misc_RNA", "precursor_RNA"}

IUPAC_NT = set("ACGTUNRYSWKMBDHV")

AA_1_TO_3 = {
    "A": "Ala", "C": "Cys", "D": "Asp", "E": "Glu", "F": "Phe", "G": "Gly",
    "H": "His", "I": "Ile", "K": "Lys", "L": "Leu", "M": "Met", "N": "Asn",
    "P": "Pro", "Q": "Gln", "R": "Arg", "S": "Ser", "T": "Thr", "V": "Val",
    "W": "Trp", "Y": "Tyr", "U": "Sec", "O": "Pyl", "*": "TERM",
}
KNOWN_AA_3 = set(AA_1_TO_3.values()) | {"Xaa", "OTHER"}


class ConversionError(Exception):
    pass


def warn(msg, warnings):
    warnings.append(msg)


def normalize_aa(raw, warnings, row_label):
    """Accepts a 1-letter or 3-letter amino acid code; returns the 3-letter
    INSDC form for the qualifier value, or None (with a warning) if
    unrecognized."""
    if not raw:
        return None
    raw = str(raw).strip()
    if len(raw) == 1:
        aa3 = AA_1_TO_3.get(raw.upper())
        if aa3:
            return aa3
    else:
        cap = raw[0].upper() + raw[1:].lower()
        if cap in KNOWN_AA_3 or raw.upper() == "TERM" or raw.upper() == "OTHER":
            return "TERM" if raw.upper() == "TERM" else ("OTHER" if raw.upper() == "OTHER" else cap)
    warn(f"Row(s) {row_label}: unrecognized amino acid code '{raw}' in Transl Except — qualifier skipped.",
         warnings)
    return None


def parse_transl_except(text, intervals, strand, codon_start, warnings, row_label):
    """Parses a 'codonNumber:aa[|codonNumber:aa...]' Transl Except cell into
    a list of ready-to-write '(pos:<loc>,aa:<AA3>)' qualifier value strings.
    Only supported for a single-interval (non-spliced) CDS — see module
    docstring.

    Unlike xlsx_to_genbank.py, this script's `intervals` hold the RAW
    entered Start/End (e.g. (17, 3) for a minus-strand row entered as
    Start > End) rather than pre-normalized ascending coordinates — they
    must be normalized to ascending here before doing any position/length
    arithmetic, or a minus-strand row's length would come out negative and
    every codon would look "out of range" (caught by testing the minus-
    strand readthrough case, which triggered exactly that)."""
    if not text:
        return []
    if len(intervals) != 1:
        warn(f"Row(s) {row_label}: Transl Except is only supported for a non-spliced "
             f"(single-interval) CDS — skipped.", warnings)
        return []
    try:
        raw_s, raw_e = float(intervals[0][0]), float(intervals[0][1])
    except (TypeError, ValueError):
        return []
    s, e = int(min(raw_s, raw_e)), int(max(raw_s, raw_e))
    length = e - s + 1
    try:
        cs = int(codon_start) if codon_start not in (None, "") else 1
    except (TypeError, ValueError):
        cs = 1

    out = []
    for chunk in str(text).split("|"):
        chunk = chunk.strip()
        if not chunk or ":" not in chunk:
            continue
        num_str, aa_str = chunk.split(":", 1)
        try:
            codon_number = int(float(num_str.strip()))
        except ValueError:
            warn(f"Row(s) {row_label}: could not parse codon number in Transl Except entry '{chunk}'.",
                 warnings)
            continue
        offset = (cs - 1) + (codon_number - 1) * 3
        if offset < 0 or offset + 3 > length:
            warn(f"Row(s) {row_label}: codon {codon_number} in Transl Except is out of range "
                 f"for this CDS — skipped.", warnings)
            continue
        if strand == "-":
            g_end = e - offset
            g_start = g_end - 2
            loc = f"complement({g_start}..{g_end})"
        else:
            g_start = s + offset
            g_end = g_start + 2
            loc = f"{g_start}..{g_end}"
        aa3 = normalize_aa(aa_str, warnings, row_label)
        if aa3 is None:
            continue
        out.append(f"(pos:{loc},aa:{aa3})")
    return out


def read_sequence_text(ws, warnings):
    """Sequence lives on its own sheet (column A, from row 2 down, one chunk
    per cell) because Excel caps a single cell at ~32,767 characters, which
    real genomes routinely exceed. Returns the full concatenated sequence
    (used for the multi-FASTA companion file on multi-segment workbooks);
    pass None if the sheet is missing."""
    chunks = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        v = row[0].value
        if v:
            chunks.append(str(v).strip())
    raw = "".join(chunks)
    if not raw:
        warn("The 'Sequence' sheet is empty — skipping coordinate range checks.", warnings)
        return None
    seq = re.sub(r"[\s\d]", "", raw).upper()
    bad = set(seq) - IUPAC_NT
    if bad:
        warn(f"Sequence contains unexpected characters: {sorted(bad)}", warnings)
    return seq


def read_sequence_length(ws, warnings):
    seq = read_sequence_text(ws, warnings)
    return None if seq is None else len(seq)


def fasta_wrap(name, seq, width=70):
    lines = [f">{name}"]
    for i in range(0, len(seq), width):
        lines.append(seq[i:i + width])
    return "\n".join(lines) + "\n"


def read_record_info(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, max_col=2):
        label_cell, value_cell = row[0], row[1]
        if label_cell.value is None:
            continue
        label = str(label_cell.value).replace("*", "").strip()
        value = value_cell.value
        value = "" if value is None else str(value).strip()
        data[label] = value
    return data


def read_features(ws):
    header_row = 2
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    required = ["Feature Key", "Start", "End"]
    for r in required:
        if r not in headers:
            raise ConversionError(f"Features sheet is missing required column '{r}'.")

    def find_col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    col_map = {
        "feature_key": find_col("Feature Key"),
        "start": find_col("Start"),
        "end": find_col("End"),
        "gene_num": find_col("Gene #"),
        "product": find_col("Product"),
        "transl_table": find_col("transl_table"),
        "codon_start": find_col("Codon Start"),
        "note": find_col("Note"),
        "other": find_col("Other Qualifiers (gb only, key=value|key=value)"),
        "partial5": find_col("Partial 5' end (Y/N)"),
        "partial3": find_col("Partial 3' end (Y/N)"),
        "exon_group": find_col("Exon Group"),
        "exception": find_col("Exception"),
        "transl_except": find_col("Transl Except"),
    }

    def get(row, key):
        col = col_map.get(key)
        if col is None:
            return ""
        v = ws.cell(row=row, column=col).value
        return "" if v is None else (str(v).strip() if not isinstance(v, (int, float)) else v)

    features = []
    for r in range(header_row + 1, ws.max_row + 1):
        fkey = get(r, "feature_key")
        if not fkey or str(fkey).strip() == "":
            continue
        note_val = get(r, "note")
        if isinstance(note_val, str) and note_val.strip().lower().startswith("example row"):
            note_val = ""
        features.append({
            "row": r,
            "feature_key": str(fkey).strip(),
            "start": get(r, "start"),
            "end": get(r, "end"),
            "gene_num": get(r, "gene_num"),
            "product": get(r, "product"),
            "transl_table": get(r, "transl_table"),
            "codon_start": get(r, "codon_start"),
            "note": note_val,
            "other": get(r, "other"),
            "partial5": str(get(r, "partial5") or "N").strip().upper() == "Y",
            "partial3": str(get(r, "partial3") or "N").strip().upper() == "Y",
            "exon_group": get(r, "exon_group"),
            "exception": get(r, "exception"),
            "transl_except": get(r, "transl_except"),
        })
    return features


def group_exons(features, warnings):
    """Merge rows sharing the same Feature Key + the same non-blank Exon
    Group value into one multi-interval feature, in sheet order. The first
    row of a group carries all qualifiers; every row in the group
    contributes one (start, end) interval. Rows with a blank Exon Group are
    left as single-interval features, unchanged from prior behavior."""
    grouped = []
    group_index = {}  # (feature_key, exon_group) -> position in `grouped`

    for feat in features:
        eg = str(feat.get("exon_group") or "").strip()
        feat["intervals"] = [(feat["start"], feat["end"])]
        feat["_exon_rows"] = [feat["row"]]
        if not eg:
            grouped.append(feat)
            continue

        gkey = (feat["feature_key"], eg)
        if gkey not in group_index:
            group_index[gkey] = len(grouped)
            grouped.append(feat)
        else:
            head = grouped[group_index[gkey]]
            try:
                head_dir = float(head["intervals"][0][0]) <= float(head["intervals"][0][1])
                this_dir = float(feat["start"]) <= float(feat["end"])
                if head_dir != this_dir:
                    warn(f"Row {feat['row']}: Exon Group '{eg}' mixes rows whose Start/End imply "
                         f"different strands — check this group's coordinates.", warnings)
            except (TypeError, ValueError):
                pass
            head["intervals"].append((feat["start"], feat["end"]))
            head["_exon_rows"].append(feat["row"])
            for field, label in (("product", "Product"), ("gene_num", "Gene #"),
                                  ("note", "Note"), ("other", "Other Qualifiers")):
                v = feat.get(field, "")
                if v not in ("", None):
                    warn(f"Row {feat['row']}: {label} on an Exon Group '{eg}' continuation row is "
                         f"ignored — only the first row of an Exon Group supplies qualifiers.", warnings)

    return grouped


def is_monotonic_ascending(intervals):
    """True if each interval's start comes strictly after the previous
    interval's end — i.e. ordinary ascending, non-wrapping order. Used to
    detect a circular-origin-spanning feature (plus strand only)."""
    for i in range(1, len(intervals)):
        try:
            prev_end = float(intervals[i - 1][1])
            this_start = float(intervals[i][0])
        except (TypeError, ValueError):
            return True  # can't tell — don't flag as a wrap
        if this_start <= prev_end:
            return False
    return True


def parse_other_qualifiers(text):
    pairs = []
    if not text:
        return pairs
    for chunk in str(text).split("|"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "=" in chunk:
            k, v = chunk.split("=", 1)
            pairs.append((k.strip(), v.strip()))
        else:
            pairs.append((chunk, None))
    return pairs


def numeric_str(v):
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return None


def build_decorated_intervals(intervals, partial5, partial3, warnings, row_label):
    """Convert raw (start, end) pairs to decorated (s, e) string pairs.
    Partial markers ('<' / '>') are applied positionally — column 1 of the
    FIRST interval gets '<', column 2 of the LAST interval gets '>' —
    matching this format's existing single-interval convention (no
    strand-based flipping; that nuance is specific to true INSDC join()
    syntax in the flat-file converter). Returns None if any interval fails
    to parse as numeric."""
    decorated = []
    for s_raw, e_raw in intervals:
        s = numeric_str(s_raw)
        e = numeric_str(e_raw)
        if s is None or e is None:
            warn(f"Row(s) {row_label}: non-numeric Start/End — feature skipped.", warnings)
            return None
        decorated.append([s, e])
    if partial5:
        decorated[0][0] = "<" + decorated[0][0]
    if partial3:
        decorated[-1][1] = ">" + decorated[-1][1]
    return decorated


def check_ranges(intervals, seq_length, warnings, row_label, key):
    if seq_length is None:
        return
    for s_raw, e_raw in intervals:
        try:
            lo = min(float(s_raw), float(e_raw))
            hi = max(float(s_raw), float(e_raw))
            if lo < 1 or hi > seq_length:
                warn(f"Row(s) {row_label}: {key} location {s_raw}..{e_raw} is out of range "
                     f"for a {seq_length}-bp sequence — check this row.", warnings)
        except (TypeError, ValueError):
            pass


def outer_bounds(intervals):
    """Overall min/max across every coordinate in a (possibly multi-interval)
    feature, used for the single-span paired gene feature of an ordinary
    (non-origin-wrapping) spliced CDS."""
    coords = []
    for s_raw, e_raw in intervals:
        try:
            coords.append(float(s_raw))
            coords.append(float(e_raw))
        except (TypeError, ValueError):
            return None, None
    if not coords:
        return None, None
    return min(coords), max(coords)


def qual_line(key, value=None, warnings=None, row=None):
    # This format is tab/newline-delimited, so a stray tab or line break
    # inside a value would silently corrupt the column structure — collapse
    # any such whitespace to single spaces instead.
    if value is not None and ("\t" in str(value) or "\n" in str(value) or "\r" in str(value)):
        if warnings is not None:
            warn(f"Row {row}: qualifier '{key}' contained a tab/newline character — replaced with a space "
                 f"(the feature table format uses tabs/newlines as delimiters).", warnings)
        value = re.sub(r"[\t\r\n]+", " ", str(value)).strip()
    if value is None or value == "":
        return f"\t\t\t{key}"
    return f"\t\t\t{key}\t{value}"


def build_feature_table(record, features, seq_length, warnings):
    name = record.get("Locus/Sequence Name", "") or "SEQ1"
    lines = [f">Feature\t{name}"]

    for feat in features:
        key = feat["feature_key"]
        if key not in VALID_FEATURE_KEYS:
            warn(f"Row {feat['row']}: unrecognized feature key '{key}' — skipped.", warnings)
            continue
        if key == "source":
            # BankIt collects organism/strain/etc. through its own web form;
            # a source row here would be unusual, but pass it through
            # generically rather than reject it outright.
            pass

        def ql(k, v=None, _row=feat["row"]):
            return qual_line(k, v, warnings, _row)

        row_label = "+".join(str(r) for r in feat["_exon_rows"])
        intervals = feat["intervals"]

        decorated = build_decorated_intervals(intervals, feat["partial5"], feat["partial3"],
                                               warnings, row_label)
        if decorated is None:
            continue
        check_ranges(intervals, seq_length, warnings, row_label, key)

        gene_num = feat["gene_num"]
        locus_tag = f"{name}_gp{int(float(gene_num))}" if gene_num not in ("", None) else None
        gp_label = f"gp{int(float(gene_num))}" if gene_num not in ("", None) else None

        if key == "CDS":
            # Paired gene feature first. Ordinarily a single span covering
            # the outer bounds only, even if the CDS itself is spliced
            # across exons — EXCEPT when the exons aren't in ascending
            # order (a circular-origin wrap, plus strand only): then the
            # gene gets the same multi-interval location as the CDS,
            # because a min/max span would incorrectly claim the gene
            # covers the entire genome in between the wrapped pieces.
            # Only checked for plus-strand groups: a normal minus-strand
            # spliced group is legitimately non-ascending in entry order
            # (see module docstring — that's the expected transcription-
            # order convention there, not a wrap), so applying this same
            # check to minus-strand groups would misfire on every ordinary
            # spliced minus-strand feature.
            try:
                first_is_plus = float(intervals[0][0]) <= float(intervals[0][1])
            except (TypeError, ValueError):
                first_is_plus = True
            origin_wrap = (len(intervals) > 1 and first_is_plus and
                           not is_monotonic_ascending(intervals))
            if origin_wrap:
                warn(f"Row(s) {row_label}: Exon Group intervals are not in ascending order — "
                     f"treating this as a circular-origin-spanning feature (plus strand). If that's "
                     f"not what you intended, check the Start/End values.", warnings)
                lines.append(f"{decorated[0][0]}\t{decorated[0][1]}\tgene")
                for s, e in decorated[1:]:
                    lines.append(f"{s}\t{e}")
                if locus_tag:
                    lines.append(ql("locus_tag", locus_tag))
                else:
                    warn(f"Row(s) {row_label}: CDS has no Gene # — gene feature written without a locus_tag.",
                         warnings)
            else:
                gmin, gmax = outer_bounds(intervals)
                if gmin is None:
                    warn(f"Row(s) {row_label}: could not compute gene span — gene feature skipped.", warnings)
                else:
                    g_s, g_e = numeric_str(gmin), numeric_str(gmax)
                    if feat["partial5"]:
                        g_s = "<" + g_s
                    if feat["partial3"]:
                        g_e = ">" + g_e
                    lines.append(f"{g_s}\t{g_e}\tgene")
                    if locus_tag:
                        lines.append(ql("locus_tag", locus_tag))
                    else:
                        warn(f"Row(s) {row_label}: CDS has no Gene # — gene feature written without a locus_tag.",
                             warnings)

            lines.append(f"{decorated[0][0]}\t{decorated[0][1]}\tCDS")
            for s, e in decorated[1:]:
                lines.append(f"{s}\t{e}")
            if not feat["product"]:
                warn(f"Row(s) {row_label}: CDS has no Product — BankIt requires one.", warnings)
            lines.append(ql("product", feat["product"] or "hypothetical protein"))
            if gp_label:
                lines.append(ql("product", gp_label))
            transl_table = feat["transl_table"] or "11"
            codon_start = feat["codon_start"] or "1"
            lines.append(ql("transl_table", transl_table))
            lines.append(ql("codon_start", codon_start))
            if feat.get("exception"):
                lines.append(ql("exception", feat["exception"]))
            single_strand = "+"
            if len(intervals) == 1:
                try:
                    if float(intervals[0][0]) > float(intervals[0][1]):
                        single_strand = "-"
                except (TypeError, ValueError):
                    pass
            for val in parse_transl_except(feat.get("transl_except"), intervals, single_strand,
                                            codon_start, warnings, row_label):
                lines.append(ql("transl_except", val))
            if feat["note"]:
                lines.append(ql("note", feat["note"]))
            for k, v in parse_other_qualifiers(feat["other"]):
                lines.append(ql(k, v))

        elif key == "gene":
            lines.append(f"{decorated[0][0]}\t{decorated[0][1]}\tgene")
            for s, e in decorated[1:]:
                lines.append(f"{s}\t{e}")
            if locus_tag:
                lines.append(ql("locus_tag", locus_tag))
            if feat["note"]:
                lines.append(ql("note", feat["note"]))
            for k, v in parse_other_qualifiers(feat["other"]):
                lines.append(ql(k, v))

        else:  # mat_peptide, tRNA/rRNA/etc., and the wider generic INSDC keys
            lines.append(f"{decorated[0][0]}\t{decorated[0][1]}\t{key}")
            for s, e in decorated[1:]:
                lines.append(f"{s}\t{e}")
            if feat["product"]:
                lines.append(ql("product", feat["product"]))
            elif key == "mat_peptide" or key in RNA_PRODUCT_KEYS:
                warn(f"Row(s) {row_label}: {key} has no Product — recommended "
                     f"(required by NCBI for mat_peptide).", warnings)
            if feat["note"]:
                lines.append(ql("note", feat["note"]))
            for k, v in parse_other_qualifiers(feat["other"]):
                lines.append(ql(k, v))

    return "\n".join(lines) + "\n"


def convert_one_segment(record, raw_features, seq_length, warnings):
    features = group_exons(raw_features, warnings)
    return build_feature_table(record, features, seq_length, warnings)


def detect_segments(wb):
    """Look for repeated "Record Info (X)"/"Sequence (X)"/"Features (X)"
    sheet triplets. Returns a list of (label, record_ws, sequence_ws,
    features_ws) tuples in first-seen order, or None if the workbook uses
    the plain single-segment sheet names ("Record Info"/"Sequence"/
    "Features")."""
    if {"Record Info", "Sequence", "Features"}.issubset(set(wb.sheetnames)):
        return None

    pattern = re.compile(r"^(Record Info|Sequence|Features)\s*\((.+)\)\s*$")
    labels = []
    seen = set()
    for name in wb.sheetnames:
        m = pattern.match(name)
        if m:
            label = m.group(2).strip()
            if label not in seen:
                seen.add(label)
                labels.append(label)

    if not labels:
        return None

    segments = []
    for label in labels:
        rec_name, seq_name, feat_name = (f"Record Info ({label})", f"Sequence ({label})",
                                          f"Features ({label})")
        missing = [n for n in (rec_name, seq_name, feat_name) if n not in wb.sheetnames]
        if missing:
            raise ConversionError(f"Segment '{label}' is missing sheet(s): {', '.join(missing)}.")
        segments.append((label, wb[rec_name], wb[seq_name], wb[feat_name]))
    return segments


def convert(xlsx_path, out_path):
    warnings = []
    wb = load_workbook(xlsx_path, data_only=True)

    segments = detect_segments(wb)
    fasta_text = None

    if segments is None:
        if "Record Info" not in wb.sheetnames:
            raise ConversionError("Workbook has no 'Record Info' sheet.")
        if "Features" not in wb.sheetnames:
            raise ConversionError("Workbook has no 'Features' sheet.")
        record = read_record_info(wb["Record Info"])
        raw_features = read_features(wb["Features"])
        seq_length = read_sequence_length(wb["Sequence"], warnings) if "Sequence" in wb.sheetnames else None
        text = convert_one_segment(record, raw_features, seq_length, warnings)
    else:
        blocks = []
        fasta_blocks = []
        for label, rec_ws, seq_ws, feat_ws in segments:
            record = read_record_info(rec_ws)
            raw_features = read_features(feat_ws)
            seq = read_sequence_text(seq_ws, warnings)
            seq_length = None if seq is None else len(seq)
            blocks.append(convert_one_segment(record, raw_features, seq_length, warnings))
            seg_name = record.get("Locus/Sequence Name", "") or label
            if seq:
                fasta_blocks.append(fasta_wrap(seg_name, seq))
        text = "\n".join(blocks)
        fasta_text = "".join(fasta_blocks)

    with open(out_path, "w", newline="\n") as f:
        f.write(text)

    fasta_path = None
    if fasta_text is not None:
        fasta_path = re.sub(r"\.(tbl\.txt|txt|tbl)$", "", out_path) + ".fsa"
        if fasta_path == out_path:
            fasta_path = out_path + ".fsa"
        with open(fasta_path, "w", newline="\n") as f:
            f.write(fasta_text)

    return text, warnings, fasta_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 xlsx_to_feature_table.py input.xlsx [output.tbl.txt]")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else re.sub(r"\.xlsx$", ".tbl.txt", xlsx_path)

    try:
        text, warnings, fasta_path = convert(xlsx_path, out_path)
    except ConversionError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    print(f"Wrote {out_path}")
    if fasta_path:
        print(f"Wrote companion multi-FASTA {fasta_path} (this is a multi-segment workbook)")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("No warnings.")


if __name__ == "__main__":
    main()
