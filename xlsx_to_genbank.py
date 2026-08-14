#!/usr/bin/env python3
"""
xlsx_to_genbank.py — convert a filled copy of GenBank_Annotation_Template.xlsx
into a spec-compliant GenBank flat file (.gb).

Usage:
    python3 xlsx_to_genbank.py input.xlsx [output.gb]

Reads the "Record Info", "Sequence", and "Features" sheets and writes a
GenBank flat file with LOCUS / DEFINITION / ACCESSION / VERSION / KEYWORDS /
SOURCE / ORGANISM / REFERENCE / COMMENT / FEATURES / ORIGIN blocks, following
the column layout used by real NCBI GenBank records (verified against a live
record, NC_012920, the NCBI feature-table spec at
https://www.ncbi.nlm.nih.gov/genbank/feature_table/, and a real published
phage genome record, PQ613263, fetched from NCBI and diffed feature-by-
feature against this converter's output).

Column conventions implemented:
  - LOCUS line: fixed-column fields (name @13, length @30-40 right-justified,
    'bp' @42-43, molecule type @48-53, topology @56-63, division @65-67,
    date @69-79).
  - FEATURES table: feature key starts at column 6, location/qualifiers start
    at column 22, lines wrapped to 79 characters with 21-space continuation
    indent.
  - ORIGIN: sequence in lowercase, 6 groups of 10 bases per line, right-
    justified 9-column position number.

Features-sheet conventions (v2, matches the feature-table converter):
  - No Strand column — Start > End means '-' strand, inferred automatically.
  - No source rows — the source feature is always built from Record Info.
  - Gene # drives a /locus_tag of "{Locus Name}_gp{N}" on both the CDS and
    its automatically-written paired gene feature (one gene per CDS,
    confirmed 1:1 against a real published record).
  - CDS always gets /codon_start (default 1) and /transl_table (default 11).
  - Only ONE /product is legal per feature in a real flat file. When a Gene #
    is given, its "gp{N}" short label is folded into /note (as "<note>; gpN"
    or just "gpN") instead of being written as a second /product — this
    matches NCBI's own behavior when it processes a raw feature-table
    submission into the published flat file (confirmed against PQ613263).
  - The first codon of every complete (non-5'-partial) CDS is always
    translated as Met, even for alternate start codons like GTG/TTG that
    would translate as Val/Leu anywhere else in the frame — standard
    NCBI/genetic-code-table-11 convention (Biopython: `.translate(cds=True)`
    implements the same rule). Missing this was a real bug caught by diffing
    against PQ613263: several minus-strand genes with a GTG start codon came
    out with a leading V instead of M until this was added.

v3 additions (for viruses beyond simple non-spliced dsDNA phage genomes):
  - mat_peptide feature key, for polyprotein viruses (coronaviruses,
    flaviviruses, picornaviruses). mat_peptide rows are independent — no
    auto-generated paired gene feature, no locus_tag unless the row
    explicitly gives one — matching real published polyprotein records.
  - "genomic RNA" added to the Molecule Type -> LOCUS mapping, for RNA
    virus genomes (LOCUS line shows "RNA"; /mol_type="genomic RNA" on the
    source feature).
  - Exon Group column: rows sharing the same Feature Key + the same
    non-blank Exon Group value are merged into one multi-interval feature
    for spliced genes in eukaryotic DNA viruses. Enter exon rows in
    transcription (5'->3') order — for a minus-strand group that means
    DESCENDING genomic order (first row = the exon at the higher
    coordinates), matching the same convention already used for a single
    minus-strand row (Start > End) and matching how NCBI's own tab-
    delimited feature table lists a spliced minus-strand feature's
    continuation lines (confirmed against NCBI's own worked example: a
    spliced tRNA-Phe). The flat file's join() operator itself is always
    written in ASCENDING genomic order regardless of strand (confirmed
    against the INSDC spec's own worked examples) — this script reverses a
    minus-strand group's entry order internally before building the join()
    string and before assembling the coding sequence, so the *input*
    convention matches the feature-table converter's while the *output*
    matches true INSDC syntax. Partial markers are placed on the true
    biological outer ends of the whole joined feature, not on every exon.
    A spliced CDS's auto-generated paired gene feature is a single span
    (outer bounds only, not joined) — matching real annotation practice
    where gene features cover introns as one range and only CDS/mRNA get
    per-exon locations. Translation assembles the coding sequence by
    concatenating raw exon regions in ascending genomic order and then
    reverse-complementing the whole assembly once for minus-strand groups —
    matching the INSDC spec's own description of complement(join(...))
    ("joins regions..., then complements the joined segments") — before
    applying the same Met-start-codon convention used for unspliced CDSs.

v4 additions (broader virus/genome coverage):
  - VALID_FEATURE_KEYS now covers the full INSDC feature key vocabulary
    (confirmed against https://www.insdc.org/submitting-standards/feature-table/),
    not just the handful this tool originally special-cased. Feature keys
    confirmed to take a /product qualifier (misc_feature, sig_peptide,
    propeptide, transit_peptide, D_segment, N_region, plus the RNA keys
    already handled) get one if the Product column is filled in; keys that
    don't take /product per spec (5'UTR, 3'UTR, repeat_region, regulatory,
    operon, stem_loop, mobile_element, etc.) fold a non-blank Product into
    /note instead, with a warning, rather than silently emitting a
    spec-invalid qualifier or silently dropping what the user typed. Two
    qualifiers that come up often with these keys are reachable through
    the existing Other Qualifiers column: `regulatory_class=promoter` (or
    terminator, ribosome_binding_site, etc.) for a `regulatory` feature,
    and `rpt_type=long_terminal_repeat` (or inverted, direct, dispersed,
    etc.) for `repeat_region` — the modern, correct way to annotate
    retroviral LTRs and viral inverted terminal repeats (INSDC dropped the
    old standalone "LTR"/"inverted_repeat" keys in favor of repeat_region +
    rpt_type). mobile_element additionally has a MANDATORY
    /mobile_element_type qualifier — also reachable via Other Qualifiers,
    and the tool does not currently warn if it's missing (a real gap; see
    TROUBLESHOOTING.md).
  - Exception, Transl Except, and Translation Override columns, for CDS
    features. Exception writes /exception="..." (e.g. "RNA editing" for
    paramyxovirus P/V/W-style edited transcripts, or "trans-splicing") and,
    unless a Translation Override is also given, suppresses the computed
    /translation — a direct extraction from the genomic sequence would not
    reflect the actual edited/spliced product, so this tool doesn't
    pretend to compute one. Transl Except writes one or more
    /transl_except=(pos:<location>,aa:<amino_acid>) qualifiers for a
    single-codon deviation from the standard translation — most commonly a
    stop-codon readthrough (alphaviruses, some plant viruses) or
    selenocysteine incorporation — and this tool DOES correctly recompute
    the translation around it (continuing past the stop rather than
    truncating there), because that's fully derivable from the genomic
    sequence plus the one stated exception. You give the codon NUMBER
    within the CDS (e.g. "142:Trp"); the genomic location is computed for
    you. Only supported for non-spliced (single-interval) CDS rows — see
    the parse_transl_except_gb docstring for why. Translation Override lets
    you supply the real protein sequence directly whenever the tool can't
    derive it (RNA-edited or trans-spliced transcripts); when given, it's
    used verbatim and no translation is computed.
  - Native multi-segment workbooks: a workbook may contain repeated sheet
    triplets named "Record Info (X)"/"Sequence (X)"/"Features (X)" (X = a
    segment label, e.g. "PB2") for genomes like influenza or bunyaviruses
    where each segment is its own GenBank record. A plain "Record Info"/
    "Sequence"/"Features" workbook (no suffix) is unchanged (single
    segment). Multi-segment output is one .gb file containing multiple
    concatenated LOCUS...// records — a standard, Biopython-`SeqIO.parse`-
    readable multi-record GenBank file.
  - Plus-strand circular-origin-spanning features: mirrors the same
    detection and gene-span fix described in xlsx_to_feature_table.py's
    v4 changelog. Minus-strand origin-spanning features remain
    unsupported (flagged with a warning), for the same reason given there.
"""
import sys
import re
import textwrap
import datetime
from openpyxl import load_workbook

LINE_WIDTH = 79
FEATURE_KEY_COL = 6      # 1-based column where the feature key starts
QUALIFIER_COL = 22       # 1-based column where location/qualifiers start
QUAL_INDENT = " " * (QUALIFIER_COL - 1)

# Full INSDC feature key vocabulary (source: official DDBJ/ENA/GenBank
# Feature Table Definition, https://www.insdc.org/submitting-standards/feature-table/).
VALID_FEATURE_KEYS = {
    "assembly_gap", "C_region", "CDS", "centromere", "D-loop", "D_segment",
    "exon", "gap", "gene", "iDNA", "intron", "J_segment", "mat_peptide",
    "misc_binding", "misc_difference", "misc_feature", "misc_recomb",
    "misc_RNA", "misc_structure", "mobile_element", "modified_base", "mRNA",
    "ncRNA", "N_region", "old_sequence", "operon", "oriT", "polyA_site",
    "precursor_RNA", "prim_transcript", "primer_bind", "propeptide",
    "protein_bind", "regulatory", "repeat_region", "rep_origin", "rRNA",
    "S_region", "sig_peptide", "source", "stem_loop", "STS", "telomere",
    "tmRNA", "transit_peptide", "tRNA", "unsure", "V_region", "V_segment",
    "variation", "3'UTR", "5'UTR",
}

# Feature keys (besides CDS/mat_peptide, handled separately) confirmed by
# the INSDC spec to take a /product qualifier.
RNA_PRODUCT_KEYS = {"tRNA", "rRNA", "tmRNA", "ncRNA", "misc_RNA", "precursor_RNA"}
OTHER_PRODUCT_OK_KEYS = {"misc_feature", "sig_peptide", "propeptide", "transit_peptide",
                          "D_segment", "N_region"}
IUPAC_NT = set("ACGTUNRYSWKMBDHV")

AA_1_TO_3 = {
    "A": "Ala", "C": "Cys", "D": "Asp", "E": "Glu", "F": "Phe", "G": "Gly",
    "H": "His", "I": "Ile", "K": "Lys", "L": "Leu", "M": "Met", "N": "Asn",
    "P": "Pro", "Q": "Gln", "R": "Arg", "S": "Ser", "T": "Thr", "V": "Val",
    "W": "Trp", "Y": "Tyr", "U": "Sec", "O": "Pyl", "*": "TERM",
}
AA_3_TO_1 = {v: k for k, v in AA_1_TO_3.items() if k not in ("U", "O", "*")}
AA_3_TO_1.update({"Sec": "U", "Pyl": "O"})
KNOWN_AA_3 = set(AA_1_TO_3.values()) | {"Xaa", "OTHER"}

MOL_TYPE_TO_LOCUS = {
    "genomic DNA": "DNA", "genomic RNA": "RNA", "mRNA": "mRNA", "rRNA": "rRNA",
    "tRNA": "tRNA", "transcribed RNA": "RNA", "other DNA": "DNA", "other RNA": "RNA",
    "viral cRNA": "cRNA",
}

CODON_TABLE = {
    'TTT': 'F', 'TTC': 'F', 'TTA': 'L', 'TTG': 'L', 'CTT': 'L', 'CTC': 'L', 'CTA': 'L', 'CTG': 'L',
    'ATT': 'I', 'ATC': 'I', 'ATA': 'I', 'ATG': 'M', 'GTT': 'V', 'GTC': 'V', 'GTA': 'V', 'GTG': 'V',
    'TCT': 'S', 'TCC': 'S', 'TCA': 'S', 'TCG': 'S', 'CCT': 'P', 'CCC': 'P', 'CCA': 'P', 'CCG': 'P',
    'ACT': 'T', 'ACC': 'T', 'ACA': 'T', 'ACG': 'T', 'GCT': 'A', 'GCC': 'A', 'GCA': 'A', 'GCG': 'A',
    'TAT': 'Y', 'TAC': 'Y', 'TAA': '*', 'TAG': '*', 'CAT': 'H', 'CAC': 'H', 'CAA': 'Q', 'CAG': 'Q',
    'AAT': 'N', 'AAC': 'N', 'AAA': 'K', 'AAG': 'K', 'GAT': 'D', 'GAC': 'D', 'GAA': 'E', 'GAG': 'E',
    'TGT': 'C', 'TGC': 'C', 'TGA': '*', 'TGG': 'W', 'CGT': 'R', 'CGC': 'R', 'CGA': 'R', 'CGG': 'R',
    'AGT': 'S', 'AGC': 'S', 'AGA': 'R', 'AGG': 'R', 'GGT': 'G', 'GGC': 'G', 'GGA': 'G', 'GGG': 'G',
}
COMPLEMENT = str.maketrans("ACGTUNRYSWKMBDHVacgtunryswkmbdhv",
                            "TGCAANYRSWMKVHDBtgcaanyrswmkvhdb")


class ConversionError(Exception):
    pass


def warn(msg, warnings):
    warnings.append(msg)


def reverse_complement(seq):
    return seq.translate(COMPLEMENT)[::-1]


START_CODONS = {"TTG", "CTG", "ATT", "ATC", "ATA", "ATG", "GTG"}  # table 11 alternate starts


def translate_cds(seq_region, codon_start, warnings, label, partial5=False, overrides=None):
    """seq_region is already strand-corrected and exon-joined (5'->3' coding
    sense), uppercase.

    Per standard GenBank/NCBI translation convention (and genetic code table 11,
    used for bacteria/archaea/phage), the FIRST codon of a complete CDS is always
    translated as Met (M), even when it's an alternate start codon like GTG or
    TTG that would translate as Val/Leu anywhere else in the reading frame. This
    only applies when codon_start == 1 and the CDS is not 5'-partial (a partial
    CDS's first codon is a mid-gene fragment, not a true start codon).

    `overrides` is an optional {codon_number (1-based): one-letter amino acid}
    map, built from the Transl Except column, for single-codon deviations
    from the standard genetic code — most commonly a stop-codon readthrough.
    When a stop codon is hit at a position with an override, translation
    substitutes the given amino acid and CONTINUES past it (rather than
    truncating there), since a readthrough product is by definition longer
    than the "default" translation would be."""
    overrides = overrides or {}
    try:
        codon_start = int(codon_start) if codon_start not in (None, "") else 1
    except ValueError:
        codon_start = 1
    trimmed = seq_region[codon_start - 1:]
    protein = []
    for i in range(0, len(trimmed) - 2, 3):
        codon_number = i // 3 + 1
        if codon_number in overrides:
            protein.append(overrides[codon_number])
            continue
        codon = trimmed[i:i + 3]
        aa = CODON_TABLE.get(codon, "X")
        if aa == "*":
            break
        protein.append(aa)
    if protein and codon_start == 1 and not partial5:
        first_codon = trimmed[0:3]
        if first_codon in START_CODONS and protein[0] != "M" and 1 not in overrides:
            protein[0] = "M"
    if not protein:
        warn(f"{label}: translation came out empty — check start/end/strand/codon_start.", warnings)
    return "".join(protein)


def normalize_aa(raw, warnings, row_label):
    """Accepts a 1-letter or 3-letter amino acid code; returns the 3-letter
    INSDC form for the qualifier value, or None (with a warning) if
    unrecognized."""
    if not raw:
        return None
    raw = str(raw).strip()
    if len(raw) == 1:
        aa3 = AA_1_TO_3.get(raw.upper())
        if aa3:
            return aa3
    else:
        if raw.upper() == "TERM":
            return "TERM"
        if raw.upper() == "OTHER":
            return "OTHER"
        cap = raw[0].upper() + raw[1:].lower()
        if cap in KNOWN_AA_3:
            return cap
    warn(f"Row(s) {row_label}: unrecognized amino acid code '{raw}' in Transl Except — qualifier skipped.",
         warnings)
    return None


def parse_transl_except_gb(text, intervals, strand, codon_start, warnings, row_label):
    """Parses a 'codonNumber:aa[|codonNumber:aa...]' Transl Except cell.
    Returns a list of (codon_number, one_letter_aa, qualifier_value_string)
    tuples. Only supported for a single-interval (non-spliced) CDS: mapping
    a codon number to a genomic location across an intron boundary would
    need to walk the exon/intron structure and hasn't been implemented or
    tested, so a spliced CDS with Transl Except gets a warning and no
    qualifier rather than risking a silently wrong location."""
    if not text:
        return []
    if len(intervals) != 1:
        warn(f"Row(s) {row_label}: Transl Except is only supported for a non-spliced "
             f"(single-interval) CDS — skipped.", warnings)
        return []
    s, e = intervals[0]
    length = e - s + 1
    try:
        cs = int(codon_start) if codon_start not in (None, "") else 1
    except (TypeError, ValueError):
        cs = 1

    out = []
    for chunk in str(text).split("|"):
        chunk = chunk.strip()
        if not chunk or ":" not in chunk:
            continue
        num_str, aa_str = chunk.split(":", 1)
        try:
            codon_number = int(float(num_str.strip()))
        except ValueError:
            warn(f"Row(s) {row_label}: could not parse codon number in Transl Except entry '{chunk}'.",
                 warnings)
            continue
        offset = (cs - 1) + (codon_number - 1) * 3
        if offset < 0 or offset + 3 > length:
            warn(f"Row(s) {row_label}: codon {codon_number} in Transl Except is out of range "
                 f"for this CDS — skipped.", warnings)
            continue
        if strand == "-":
            g_end = e - offset
            g_start = g_end - 2
            loc = f"complement({g_start}..{g_end})"
        else:
            g_start = s + offset
            g_end = g_start + 2
            loc = f"{g_start}..{g_end}"
        aa3 = normalize_aa(aa_str, warnings, row_label)
        if aa3 is None:
            continue
        one_letter = AA_3_TO_1.get(aa3)
        out.append((codon_number, one_letter, f"(pos:{loc},aa:{aa3})"))
    return out


def wrap_header_field(label, value, width=LINE_WIDTH):
    """DEFINITION/ACCESSION/SOURCE-style header lines: 12-char label field,
    value starting col 13, continuation lines indented 12 spaces."""
    label_field = (label + " " * 12)[:12]
    indent = " " * 12
    wrapped = textwrap.wrap(value, width=width - 12) or [""]
    lines = [label_field + wrapped[0]]
    for cont in wrapped[1:]:
        lines.append(indent + cont)
    return lines


def format_locus_line(name, length, mol_type_label, topology, division, date_str):
    if len(name) > 16:
        name = name[:16]
    mol_type = MOL_TYPE_TO_LOCUS.get(mol_type_label, "DNA")
    name_f = name.ljust(16)
    length_f = str(length).rjust(11)
    moltype_f = mol_type[:6].ljust(6)
    topo_f = (topology or "linear")[:8].ljust(8)
    div_f = (division or "UNA")[:3].ljust(3)
    line = ("LOCUS" + " " * 7 + name_f + " " + length_f + " " + "bp" + " " +
            " " * 3 + moltype_f + "  " + topo_f + " " + div_f + " " + date_str)
    return line


def format_qualifier(key, value, is_translation=False):
    """Return a list of lines (each <= LINE_WIDTH chars) for a /key="value" qualifier."""
    if value is None or value == "":
        tag = f"/{key}"
    else:
        # INSDC spec: embedded double quotes in a free-text qualifier value
        # must be escaped by doubling them, e.g. /note="He said ""hi""".
        safe_value = str(value).replace('"', '""')
        tag = f'/{key}="{safe_value}"'
    if is_translation:
        return wrap_translation(tag)
    full_first_avail = LINE_WIDTH - len(QUAL_INDENT)
    if len(tag) <= full_first_avail:
        return [QUAL_INDENT + tag]
    # word-wrap on spaces, keeping the /key="  intact with the first word
    wrapped = textwrap.wrap(tag, width=full_first_avail, break_long_words=False,
                             break_on_hyphens=False)
    return [QUAL_INDENT + w for w in wrapped]


def wrap_translation(tag):
    """Hard-wrap /translation="SEQ...' at fixed column width, matching NCBI style:
    first line fits after the opening tag, continuation lines are exactly
    (LINE_WIDTH - indent) characters."""
    avail = LINE_WIDTH - len(QUAL_INDENT)
    if len(tag) <= avail:
        return [QUAL_INDENT + tag]
    prefix_len = len('/translation="')
    first_chunk_len = avail - prefix_len
    header = tag[:prefix_len]
    rest = tag[prefix_len:]
    closing = rest.endswith('"')
    body = rest[:-1] if closing else rest
    lines = []
    first = body[:first_chunk_len]
    body = body[first_chunk_len:]
    lines.append(QUAL_INDENT + header + first)
    while body:
        chunk = body[:avail]
        body = body[avail:]
        if not body and closing:
            chunk = chunk + '"'
        lines.append(QUAL_INDENT + chunk)
    return lines


def format_feature_line(key, location):
    key_field = key.ljust(QUALIFIER_COL - FEATURE_KEY_COL - 1)  # pad to col 20 (15 wide)
    prefix = " " * (FEATURE_KEY_COL - 1) + key_field + " "
    avail = LINE_WIDTH - len(prefix)
    lines = []
    if len(location) <= avail:
        lines.append(prefix + location)
    else:
        wrapped = textwrap.wrap(location, width=avail, break_long_words=True,
                                 break_on_hyphens=False)
        lines.append(prefix + wrapped[0])
        for cont in wrapped[1:]:
            lines.append(QUAL_INDENT + cont)
    return lines


def build_location_multi(intervals, strand, partial5, partial3):
    """intervals: list of (start, end) tuples in ENTRY order — the order the
    user listed Exon Group rows in the sheet. By convention (matching the
    tab-delimited feature table's own multi-interval convention — confirmed
    against NCBI's own worked example at
    https://www.ncbi.nlm.nih.gov/genbank/feature_table/: a spliced
    tRNA-Phe whose tab-table lines read "4626 4590" then "4570 4535"
    becomes complement(join(4535..4570,4590..4626)) in the resulting flat
    file), rows for a MINUS-strand Exon Group are entered in transcription
    order, which is DESCENDING genomic order for minus strand — the
    opposite of the flat file's own join() argument order. The INSDC
    feature table spec always writes join() arguments in ASCENDING genomic
    order regardless of strand (confirmed against two further worked
    examples in the official spec: complement(join(2691..4571,4918..5163))
    and complement(join(21..349,567..795))) — complementing happens to the
    joined-and-concatenated whole, not to each exon individually. So a
    minus-strand group's entry order is reversed here before the join() is
    built; a plus-strand group's entry order is already ascending.

    An earlier version of this function treated entry order as if it were
    already the correct join() order for minus strand and complemented each
    exon in place — this silently produced a location string whose parts
    were in the wrong order relative to the true INSDC convention. Caught
    by round-tripping a hand-built two-exon minus-strand test case through
    both this code and Biopython's GenBank parser, then checking both
    against the spec's own worked examples above; the corrected version
    here reproduces the spec's exact example ordering and its translated
    protein extracts identically via Biopython.

    Partial markers: per spec 3.4.2 ("complement(34..126) starts at the
    base complementary to 126"), the biological 5' end of a minus-strand
    feature is its LARGEST coordinate and the 3' end is its SMALLEST — so
    on the (now-ascending) list, '>' (5'-partial) goes on the LAST
    interval's end, and '<' (3'-partial) goes on the FIRST interval's
    start. Plus strand keeps the intuitive placement (5' = first/smallest).
    """
    ascending = list(reversed(intervals)) if strand == "-" else list(intervals)
    parts = [[str(int(s)), str(int(e))] for s, e in ascending]

    if strand == "-":
        if partial5:
            parts[-1][1] = ">" + parts[-1][1]
        if partial3:
            parts[0][0] = "<" + parts[0][0]
    else:
        if partial5:
            parts[0][0] = "<" + parts[0][0]
        if partial3:
            parts[-1][1] = ">" + parts[-1][1]

    range_strs = [f"{s}..{e}" for s, e in parts]
    loc = range_strs[0] if len(range_strs) == 1 else "join(" + ",".join(range_strs) + ")"
    if strand == "-":
        loc = f"complement({loc})"
    return loc


def build_coding_sequence(seq, intervals, strand):
    """intervals in ENTRY order (see build_location_multi's docstring for
    why entry order is descending-genomic for minus-strand groups). Per the
    INSDC spec's own description of complement(join(...)) ("joins regions
    ..., then complements the joined segments"): concatenate the RAW
    (not-yet-complemented) regions in ascending genomic order, then
    reverse-complement the whole assembled string ONCE — not each exon
    individually. (Reverse-complementing each exon separately in entry
    order, as an earlier version of this function did, silently produces
    the wrong protein for any minus-strand group with more than one exon —
    it happens to be a no-op for single-exon features, which is why the
    non-spliced case was unaffected.)"""
    ascending = list(reversed(intervals)) if strand == "-" else list(intervals)
    raw = "".join(seq[int(s) - 1:int(e)] for s, e in ascending)
    return reverse_complement(raw) if strand == "-" else raw


def read_record_info(ws):
    data = {}
    for row in ws.iter_rows(min_row=2, max_col=2):
        label_cell, value_cell = row[0], row[1]
        if label_cell.value is None:
            continue
        label = str(label_cell.value).replace("*", "").strip()
        value = value_cell.value
        value = "" if value is None else str(value).strip()
        data[label] = value
    return data


def read_features(ws):
    """Reads the v2/v3 Features sheet schema (no explicit Strand column —
    strand is inferred from Start/End order; no source rows — source always
    comes from Record Info; optional Exon Group column for spliced
    features)."""
    header_row = 2
    headers = {}
    for cell in ws[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    required = ["Feature Key", "Start", "End"]
    for r in required:
        if r not in headers:
            raise ConversionError(f"Features sheet is missing required column '{r}'.")

    def get(row, name):
        col = headers.get(name)
        if col is None:
            return ""
        v = ws.cell(row=row, column=col).value
        return "" if v is None else (str(v).strip() if not isinstance(v, (int, float)) else v)

    features = []
    for r in range(header_row + 1, ws.max_row + 1):
        fkey = get(r, "Feature Key")
        if not fkey or str(fkey).strip() == "":
            continue
        note_val = get(r, "Note")
        if isinstance(note_val, str) and note_val.strip().lower().startswith("example row"):
            note_val = ""

        raw_start = get(r, "Start")
        raw_end = get(r, "End")
        strand = "+"
        start, end = raw_start, raw_end
        try:
            s_num, e_num = float(raw_start), float(raw_end)
            if s_num > e_num:
                strand = "-"
                start, end = raw_end, raw_start
        except (TypeError, ValueError):
            pass

        gene_num = get(r, "Gene #")

        rec = {
            "row": r,
            "feature_key": str(fkey).strip(),
            "start": start,
            "end": end,
            "strand": strand,
            "partial5": str(get(r, "Partial 5' end (Y/N)") or "N").strip().upper() == "Y",
            "partial3": str(get(r, "Partial 3' end (Y/N)") or "N").strip().upper() == "Y",
            "gene_num": gene_num,
            "product": get(r, "Product"),
            "protein_id": get(r, "Protein ID (gb only)"),
            "transl_table": get(r, "transl_table"),
            "codon_start": get(r, "Codon Start"),
            "note": note_val,
            "db_xref": get(r, "db_xref (gb only)"),
            "other": get(r, "Other Qualifiers (gb only, key=value|key=value)"),
            "exon_group": get(r, "Exon Group"),
            "exception": get(r, "Exception"),
            "transl_except": get(r, "Transl Except"),
            "translation_override": get(r, "Translation Override"),
        }
        features.append(rec)
    return features


def group_exons(features, warnings):
    """Merge rows sharing the same Feature Key + the same non-blank Exon
    Group value into one multi-interval feature, in sheet order. All rows in
    a group must share the same strand (checked here); the first row
    supplies every qualifier. Non-grouped rows are left untouched, now
    carrying a single-element "intervals" list for a uniform downstream
    interface."""
    grouped = []
    group_index = {}

    for feat in features:
        eg = str(feat.get("exon_group") or "").strip()
        feat["intervals"] = [(feat["start"], feat["end"])]
        feat["_exon_rows"] = [feat["row"]]
        if not eg:
            grouped.append(feat)
            continue

        gkey = (feat["feature_key"], eg)
        if gkey not in group_index:
            group_index[gkey] = len(grouped)
            grouped.append(feat)
        else:
            head = grouped[group_index[gkey]]
            if feat["strand"] != head["strand"]:
                warn(f"Row {feat['row']}: Exon Group '{eg}' mixes + and - strand rows — "
                     f"using the first row's strand ({head['strand']}) for the whole feature; "
                     f"check this group's Start/End values.", warnings)
            head["intervals"].append((feat["start"], feat["end"]))
            head["_exon_rows"].append(feat["row"])
            for field, label in (("product", "Product"), ("gene_num", "Gene #"), ("note", "Note"),
                                  ("protein_id", "Protein ID"), ("db_xref", "db_xref"),
                                  ("other", "Other Qualifiers"), ("exception", "Exception"),
                                  ("transl_except", "Transl Except"),
                                  ("translation_override", "Translation Override")):
                v = feat.get(field, "")
                if v not in ("", None):
                    warn(f"Row {feat['row']}: {label} on an Exon Group '{eg}' continuation row is "
                         f"ignored — only the first row of an Exon Group supplies qualifiers.", warnings)

    return grouped


def is_monotonic_ascending(intervals):
    """True if each interval's start comes strictly after the previous
    interval's end — i.e. ordinary ascending, non-wrapping order. Used to
    detect a circular-origin-spanning feature (plus strand only)."""
    for i in range(1, len(intervals)):
        try:
            prev_end = float(intervals[i - 1][1])
            this_start = float(intervals[i][0])
        except (TypeError, ValueError):
            return True  # can't tell — don't flag as a wrap
        if this_start <= prev_end:
            return False
    return True


def parse_other_qualifiers(text):
    pairs = []
    if not text:
        return pairs
    for chunk in str(text).split("|"):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "=" in chunk:
            k, v = chunk.split("=", 1)
            pairs.append((k.strip(), v.strip()))
        else:
            pairs.append((chunk, None))
    return pairs


def read_sequence_ws(ws, warnings):
    """Sequence lives on its own sheet (column A, from row 2 down, one chunk
    per cell) because Excel caps a single cell at ~32,767 characters, which
    real genomes routinely exceed."""
    chunks = []
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        v = row[0].value
        if v:
            chunks.append(str(v).strip())
    raw = "".join(chunks)
    if not raw:
        raise ConversionError("The 'Sequence' sheet has no sequence data in column A (starting row 2).")
    seq = re.sub(r"[\s\d]", "", raw).upper()
    bad = set(seq) - IUPAC_NT
    if bad:
        warn(f"Sequence contains unexpected characters (kept as-is): {sorted(bad)}", warnings)
    if not seq:
        raise ConversionError("The 'Sequence' sheet has no valid nucleotide characters.")
    return seq


def read_sequence_sheet(wb, warnings):
    if "Sequence" not in wb.sheetnames:
        raise ConversionError("Workbook has no 'Sequence' sheet.")
    return read_sequence_ws(wb["Sequence"], warnings)


def outer_bounds(intervals):
    coords = []
    for s, e in intervals:
        coords.append(float(s))
        coords.append(float(e))
    return min(coords), max(coords)


def build_genbank(record, features, seq, warnings):
    name = record.get("Locus/Sequence Name", "") or "SEQ1"
    definition = record.get("Definition", "") or "."
    organism = record.get("Organism", "") or "Unclassified organism"
    mol_type_label = record.get("Molecule Type", "") or "genomic DNA"
    topology = (record.get("Topology", "") or "linear").lower()
    division_raw = record.get("Division", "") or "UNA - unannotated"
    division = division_raw.split(" - ")[0].strip()[:3].upper()
    accession = record.get("Accession", "") or name
    strain = record.get("Strain", "")
    isolate = record.get("Isolate", "")
    coll_date = record.get("Collection Date", "")
    country = record.get("Country", "")
    comment = record.get("Comment", "")

    length = len(seq)
    today = datetime.date.today().strftime("%d-%b-%Y").upper()

    lines = []
    lines.append(format_locus_line(name, length, mol_type_label, topology, division, today))
    lines.extend(wrap_header_field("DEFINITION", definition))
    lines.extend(wrap_header_field("ACCESSION", accession))
    lines.append(("VERSION" + " " * 12)[:12] + f"{accession}.1")
    lines.append(("KEYWORDS" + " " * 12)[:12] + ".")
    lines.extend(wrap_header_field("SOURCE", organism))
    lines.extend(wrap_header_field("  ORGANISM", organism))
    lines.append(" " * 12 + "Unclassified.")
    lines.append(("REFERENCE" + " " * 12)[:12] + f"1  (bases 1 to {length})")
    lines.extend(wrap_header_field("  AUTHORS", "."))
    lines.extend(wrap_header_field("  TITLE", "Direct Submission"))
    lines.extend(wrap_header_field(
        "  JOURNAL",
        f"Submitted ({today}) to GenBank. Generated with a spreadsheet-to-GenBank "
        f"conversion pipeline; replace this reference with a real citation before submission."))
    if comment:
        lines.extend(wrap_header_field("COMMENT", comment))

    lines.append("FEATURES             Location/Qualifiers")

    # The v2/v3 Features sheet has no 'source' rows (BankIt collects that
    # separately) — always build the source feature from Record Info.
    auto_source = {
        "row": None, "feature_key": "source", "intervals": [(1, length)], "strand": "+",
        "partial5": False, "partial3": False, "gene_num": "", "product": "", "protein_id": "",
        "transl_table": "", "codon_start": "", "note": "", "db_xref": "", "other": "",
    }

    # Every non-spliced CDS in a real GenBank record has a paired 'gene'
    # feature (same location) written immediately before it, carrying just
    # the locus_tag — confirmed against the published record this converter
    # was checked against (82 CDS, 82 gene, one-to-one). A spliced CDS's
    # paired gene feature is a single span over the outer bounds only (real
    # annotation practice: gene features cover introns as one range; only
    # CDS/mRNA get per-exon join() locations). mat_peptide rows do NOT get
    # an auto-generated gene feature (they're independent, matching real
    # polyprotein records).
    expanded = []
    for feat in features:
        if feat["feature_key"] == "CDS":
            intervals = feat["intervals"]
            gene_feat = dict(feat)
            gene_feat["feature_key"] = "gene"
            # Ordinarily the paired gene feature is a single span over the
            # outer bounds only, even if the CDS itself is spliced across
            # exons. EXCEPTION: if the exons aren't in ascending order for
            # a plus-strand CDS, that's a circular-origin-spanning feature
            # (e.g. a geminivirus Rep gene wrapping from near the end of a
            # circular sequence back to position 1 — the INSDC spec's own
            # worked example of this is join(2004..2195,3..20)) and a
            # min/max span would be wrong: it would claim the gene covers
            # the entire genome in between the two wrapped pieces. In that
            # case the gene gets the same multi-interval join() as the CDS.
            # Minus-strand origin wraps aren't detected/handled here — the
            # correct INSDC representation needs the alternate
            # join(complement(...),complement(...)) syntax, not implemented.
            origin_wrap = (len(intervals) > 1 and feat["strand"] == "+" and
                           not is_monotonic_ascending(intervals))
            if origin_wrap:
                warn(f"Row {feat['row']}: Exon Group intervals are not in ascending order — "
                     f"treating this as a circular-origin-spanning feature (plus strand) and giving "
                     f"the paired gene feature the same join() as the CDS. If that's not what you "
                     f"intended, check the Start/End values.", warnings)
                gene_feat["intervals"] = list(intervals)
            else:
                gmin, gmax = outer_bounds(intervals)
                gene_feat["intervals"] = [(gmin, gmax)]
            gene_feat["product"] = ""
            gene_feat["protein_id"] = ""
            gene_feat["transl_table"] = ""
            gene_feat["codon_start"] = ""
            gene_feat["note"] = ""
            gene_feat["db_xref"] = ""
            gene_feat["other"] = ""
            gene_feat["exception"] = ""
            gene_feat["transl_except"] = ""
            gene_feat["translation_override"] = ""
            if not feat.get("gene_num"):
                warn(f"Row {feat['row']}: CDS has no Gene # — paired gene feature written without a locus_tag.",
                     warnings)
            expanded.append(gene_feat)
        expanded.append(feat)

    ordered = [auto_source] + expanded

    for feat in ordered:
        key = feat["feature_key"]
        row_label = "+".join(str(r) for r in feat.get("_exon_rows", [feat["row"]]))
        if key == "source":
            pass  # always valid, built internally
        elif key not in VALID_FEATURE_KEYS:
            warn(f"Row {row_label}: unrecognized feature key '{key}' — skipped.", warnings)
            continue

        intervals = feat["intervals"]
        try:
            norm_intervals = [(int(float(s)), int(float(e))) for s, e in intervals]
        except (TypeError, ValueError):
            warn(f"Row {row_label}: {key} has non-numeric Start/End — skipped.", warnings)
            continue

        for s, e in norm_intervals:
            if s < 1 or e > length or s > e:
                warn(f"Row {row_label}: {key} location {s}..{e} is out of range for a "
                     f"{length}-bp sequence, or Start > End — check this row.", warnings)

        loc = build_location_multi(norm_intervals, feat["strand"], feat["partial5"], feat["partial3"])
        lines.extend(format_feature_line(key, loc))

        region = build_coding_sequence(seq, norm_intervals, feat["strand"])

        gene_num = feat.get("gene_num", "")
        gp_label = None
        locus_tag = None
        if gene_num not in ("", None):
            try:
                n = int(float(gene_num))
                locus_tag = f"{name}_gp{n}"
                gp_label = f"gp{n}"
            except (TypeError, ValueError):
                locus_tag = None

        quals = []
        if key == "source":
            quals.append(("organism", organism))
            quals.append(("mol_type", mol_type_label))
            if strain:
                quals.append(("strain", strain))
            if isolate:
                quals.append(("isolate", isolate))
            if country:
                quals.append(("country", country))
            if coll_date:
                quals.append(("collection_date", coll_date))
        else:
            if locus_tag:
                quals.append(("locus_tag", locus_tag))
            # A valid flat file has exactly one /product per feature. In the
            # raw feature-table format a second "product" line (the gp#
            # label) is legal shorthand, but NCBI's own processing folds it
            # into /note when generating the flatfile — confirmed against a
            # real published record (PQ613263): /note="<original note>; gp1"
            # or just /note="gp1" when there was no original note. Match
            # that here rather than emitting a second /product qualifier.
            note_text = feat["note"] or ""
            if key == "CDS" and gp_label:
                note_text = f"{note_text}; {gp_label}" if note_text else gp_label
            transl_except_entries = []
            if key == "CDS":
                # Qualifier order (note, exception, codon_start, transl_table,
                # transl_except, product, protein_id) matches the order used
                # in the real published record this converter was validated
                # against (PQ613263), with the v4 exception/transl_except
                # qualifiers slotted in alongside their closest relatives.
                if note_text:
                    quals.append(("note", note_text))
                if feat.get("exception"):
                    quals.append(("exception", feat["exception"]))
                tt = feat["transl_table"] or "11"
                cs = feat["codon_start"] or "1"
                if not feat["product"]:
                    warn(f"Row {row_label}: CDS has no Product — GenBank/BankIt requires one; "
                         f"add a product before submitting.", warnings)
                quals.append(("codon_start", cs, True))    # True = no quotes (numeric)
                quals.append(("transl_table", tt, True))
                transl_except_entries = parse_transl_except_gb(
                    feat.get("transl_except"), norm_intervals, feat["strand"], cs, warnings, row_label)
                for _, _, qual_val in transl_except_entries:
                    quals.append(("transl_except", qual_val))
                quals.append(("product", feat["product"] or "hypothetical protein"))
                if feat["protein_id"]:
                    quals.append(("protein_id", feat["protein_id"]))
            elif key in RNA_PRODUCT_KEYS or key == "mat_peptide":
                # product before note, matching the published record's order
                if feat["product"]:
                    quals.append(("product", feat["product"]))
                elif key == "mat_peptide":
                    warn(f"Row {row_label}: mat_peptide has no Product — NCBI requires one.", warnings)
                if note_text:
                    quals.append(("note", note_text))
            elif key == "gene":
                # Real gene features never carry /product.
                if note_text:
                    quals.append(("note", note_text))
            else:
                # The wider INSDC feature key vocabulary (5'UTR, repeat_region,
                # regulatory, operon, stem_loop, etc.). Only the keys
                # confirmed by the spec to take /product get one; anything
                # else with a Product entry gets it folded into /note instead
                # of emitting a non-standard qualifier.
                if feat["product"] and key in OTHER_PRODUCT_OK_KEYS:
                    quals.append(("product", feat["product"]))
                elif feat["product"]:
                    note_text = f"{note_text}; {feat['product']}" if note_text else feat["product"]
                    warn(f"Row {row_label}: '{key}' doesn't take a /product qualifier per the INSDC "
                         f"spec — the Product column value was folded into Note instead.", warnings)
                if note_text:
                    quals.append(("note", note_text))
            if feat["db_xref"]:
                quals.append(("db_xref", feat["db_xref"]))

        for extra_k, extra_v in parse_other_qualifiers(feat["other"]):
            quals.append((extra_k, extra_v))

        for q in quals:
            if len(q) == 3 and q[2] is True:
                k, v, _ = q
                lines.append(QUAL_INDENT + f"/{k}={v}")
            else:
                k, v = q[0], q[1]
                lines.extend(format_qualifier(k, v))

        if key == "CDS":
            override = (feat.get("translation_override") or "").strip()
            if override:
                protein = re.sub(r"\s+", "", override).upper()
            elif feat.get("exception"):
                protein = ""
                warn(f"Row {row_label}: CDS has /exception set — skipping the computed /translation "
                     f"(a direct extraction from the genomic sequence wouldn't reflect the actual "
                     f"edited/spliced product). Supply a Translation Override if you know the real "
                     f"protein sequence.", warnings)
            else:
                overrides = {num: one for num, one, _ in transl_except_entries if one}
                protein = translate_cds(region, feat["codon_start"] or 1, warnings, f"Row {row_label} CDS",
                                         partial5=feat.get("partial5", False), overrides=overrides)
            if protein:
                lines.extend(format_qualifier("translation", protein, is_translation=True))

    lines.append("ORIGIN")
    for i in range(0, length, 60):
        pos = i + 1
        chunk = seq[i:i + 60].lower()
        groups = [chunk[g:g + 10] for g in range(0, len(chunk), 10)]
        lines.append(f"{pos:>9} " + " ".join(groups))
    lines.append("//")

    return "\n".join(lines) + "\n"


def convert_one_segment(record, raw_features, seq, warnings):
    features = group_exons(raw_features, warnings)
    return build_genbank(record, features, seq, warnings)


def detect_segments(wb):
    """Look for repeated "Record Info (X)"/"Sequence (X)"/"Features (X)"
    sheet triplets, mirroring xlsx_to_feature_table.py's detect_segments().
    Returns a list of (label, record_ws, sequence_ws, features_ws) tuples in
    first-seen order, or None for a plain single-segment workbook."""
    if {"Record Info", "Sequence", "Features"}.issubset(set(wb.sheetnames)):
        return None

    pattern = re.compile(r"^(Record Info|Sequence|Features)\s*\((.+)\)\s*$")
    labels = []
    seen = set()
    for sheet_name in wb.sheetnames:
        m = pattern.match(sheet_name)
        if m:
            label = m.group(2).strip()
            if label not in seen:
                seen.add(label)
                labels.append(label)

    if not labels:
        return None

    segments = []
    for label in labels:
        rec_name, seq_name, feat_name = (f"Record Info ({label})", f"Sequence ({label})",
                                          f"Features ({label})")
        missing = [n for n in (rec_name, seq_name, feat_name) if n not in wb.sheetnames]
        if missing:
            raise ConversionError(f"Segment '{label}' is missing sheet(s): {', '.join(missing)}.")
        segments.append((label, wb[rec_name], wb[seq_name], wb[feat_name]))
    return segments


def convert(xlsx_path, out_path):
    warnings = []
    wb = load_workbook(xlsx_path, data_only=True)

    segments = detect_segments(wb)

    if segments is None:
        if "Record Info" not in wb.sheetnames:
            raise ConversionError("Workbook has no 'Record Info' sheet.")
        if "Features" not in wb.sheetnames:
            raise ConversionError("Workbook has no 'Features' sheet.")
        record = read_record_info(wb["Record Info"])
        raw_features = read_features(wb["Features"])
        seq = read_sequence_sheet(wb, warnings)
        gb_text = convert_one_segment(record, raw_features, seq, warnings)
    else:
        blocks = []
        for label, rec_ws, seq_ws, feat_ws in segments:
            record = read_record_info(rec_ws)
            raw_features = read_features(feat_ws)
            seq = read_sequence_ws(seq_ws, warnings)
            blocks.append(convert_one_segment(record, raw_features, seq, warnings))
        gb_text = "\n".join(blocks)

    with open(out_path, "w") as f:
        f.write(gb_text)

    return gb_text, warnings


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 xlsx_to_genbank.py input.xlsx [output.gb]")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) > 2 else re.sub(r"\.xlsx$", ".gb", xlsx_path)

    try:
        gb_text, warnings = convert(xlsx_path, out_path)
    except ConversionError as e:
        print(f"ERROR: {e}")
        sys.exit(1)

    print(f"Wrote {out_path}")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("No warnings.")


if __name__ == "__main__":
    main()
